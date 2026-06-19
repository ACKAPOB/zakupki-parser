#!/usr/bin/env python3
"""
Парсер закупок для сервера
- Если указаны ИНН в команде (-i) - используем их
- Если ИНН не указаны - берем из inn_list.txt
"""

import sys
import os
import re
from datetime import datetime, date, timedelta
import argparse
import json
import time

try:
    import requests
    from bs4 import BeautifulSoup
    import openpyxl
except ImportError as e:
    print(f"Ошибка: не установлен модуль {e.name}")
    sys.exit(1)

# Настройки
HEADERS = {
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:97.0) Gecko/20100101 Firefox/97.0",
}

BASE_URL = "https://zakupki.gov.ru"
SEARCH_URL = "https://zakupki.gov.ru/epz/order/extendedsearch/results.html"

BAD_WORDS = []  # Пустой список - ничего не фильтруем


def load_inn_from_file():
    """Загрузка ИНН из файла inn_list.txt"""
    inn_dict = {}
    try:
        with open("inn_list.txt", 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    parts = line.split('|')
                    if len(parts) >= 2:
                        inn_dict[parts[0]] = parts[1]
                    elif len(parts) == 1:
                        inn_dict[parts[0]] = parts[0]
        print(f"Загружено {len(inn_dict)} ИНН из inn_list.txt")
        return inn_dict
    except FileNotFoundError:
        print("Файл inn_list.txt не найден")
        return {}
    except Exception as e:
        print(f"Ошибка чтения inn_list.txt: {e}")
        return {}


def parse_arguments():
    parser = argparse.ArgumentParser(description="Парсер закупок ЕИС")
    parser.add_argument("-i", "--inn", nargs="+", help="ИНН (если указаны - используем их, иначе берем из inn_list.txt)")
    parser.add_argument("-y", "--year", type=int, help="Год поиска")
    parser.add_argument("-m", "--month", type=int, help="Месяц поиска (1-12)")
    parser.add_argument("--date-from", help="Дата начала (ДД.ММ.ГГГГ)")
    parser.add_argument("--date-to", help="Дата окончания (ДД.ММ.ГГГГ)")
    parser.add_argument("--days", type=int, default=30, help="Количество дней (если не указаны год/месяц)")
    parser.add_argument("--page-size", type=int, default=50, help="Записей на страницу")
    parser.add_argument("--max-pages", type=int, default=50, help="Максимум страниц")
    parser.add_argument("--delay", type=float, default=0.5, help="Задержка между запросами")
    parser.add_argument("--debug", action="store_true", help="Режим отладки")
    return parser.parse_args()


def get_output_filename():
    """Генерация имени файла в папке output"""
    now = datetime.now()
    date_str = now.strftime("%d.%m.%Y")
    
    # Создаем папку output если её нет
    os.makedirs("output", exist_ok=True)
    
    pattern = f"zakupki_{date_str}_"
    existing = [f for f in os.listdir("output") if f.startswith(pattern) and f.endswith('.xlsx')]
    next_num = len(existing) + 1
    return f"output/zakupki_{date_str}_{next_num:02d}.xlsx"


def get_date_range(args):
    """Определение диапазона дат"""
    now = datetime.now().date()
    
    # Если указаны явные даты
    if args.date_from and args.date_to:
        date_from = datetime.strptime(args.date_from, "%d.%m.%Y").date()
        date_to = datetime.strptime(args.date_to, "%d.%m.%Y").date()
        print(f"Период: {date_from} - {date_to}")
        return date_from, date_to
    
    # Если указан год и месяц
    if args.year and args.month:
        date_from = date(args.year, args.month, 1)
        if args.month == 12:
            date_to = date(args.year, 12, 31)
        else:
            date_to = date(args.year, args.month + 1, 1) - timedelta(days=1)
        print(f"Период: {date_from} - {date_to}")
        return date_from, date_to
    
    # Если указан только год
    if args.year:
        date_from = date(args.year, 1, 1)
        date_to = date(args.year, 12, 31)
        print(f"Период: {date_from} - {date_to}")
        return date_from, date_to
    
    # По умолчанию - последние N дней
    date_to = now
    date_from = now - timedelta(days=args.days)
    print(f"Период (последние {args.days} дней): {date_from} - {date_to}")
    return date_from, date_to


def build_search_params(inn, date_from, date_to, page, page_size):
    """Формирование параметров запроса"""
    return {
        "searchString": str(inn),
        "morphology": "on",
        "updateDateFrom": date_from.strftime("%d.%m.%Y"),
        "updateDateTo": date_to.strftime("%d.%m.%Y"),
        "sortBy": "ID",
        "sortDirection": "false",
        "recordsPerPage": f"_{page_size}",
        "pageNumber": str(page),
        "fz44": "on",
        "fz223": "on",
        "af": "on",
        "ca": "on",
        "pc": "on",
        "pa": "on",
        "showLotsInfoHidden": "false",
        "OrderPlacementSmallBusinessSubject": "on",
        "OrderPlacementRnpData": "on",
        "OrderPlacementExecutionRequirement": "on",
        "priceContractAdvantages44IdNameHidden": "%7B%7D",
        "priceContractAdvantages94IdNameHidden": "%7B%7D",
        "currencyIdGeneral": "-1",
        "selectedSubjectsIdNameHidden": "%7B%7D",
        "contractPriceCurrencyId": "-1",
        "budgetLevelIdNameHidden": "%7B%7D",
        "nonBudgetTypesIdNameHidden": "%7B%7D",
        "orderPlacement94_0": "0",
        "orderPlacement94_1": "0",
        "orderPlacement94_2": "0",
    }


def parse_price(price_text):
    """Преобразование цены в число"""
    if not price_text:
        return 0.0
    cleaned = re.sub(r"[₽\s]", "", price_text)
    cleaned = cleaned.replace(",", ".").replace(chr(160), "")
    try:
        return float(cleaned)
    except:
        return 0.0


def parse_date(date_text):
    """Преобразование даты"""
    if not date_text:
        return None
    try:
        return datetime.strptime(date_text.strip(), "%d.%m.%Y").date()
    except:
        return None


def extract_purchase_info(block):
    """Извлечение информации о закупке из HTML-блока"""
    try:
        record = {}
        
        # ФЗ и тип
        header_title = block.find("div", class_="registry-entry__header-top__title")
        if header_title:
            values = [e.strip() for e in header_title.text.split("\n") if e.strip()]
            if len(values) >= 2:
                record["fz"] = values[0]
                record["type"] = values[1]
        
        # Номер и ссылка
        number_block = block.find("div", class_="registry-entry__header-mid__number")
        if number_block:
            link_tag = number_block.find("a")
            if link_tag:
                href = link_tag.get("href", "")
                if record.get("fz") == "44-ФЗ":
                    record["link"] = BASE_URL + href
                else:
                    record["link"] = href
                
                number_text = link_tag.text.strip()
                match = re.search(r"№\s*([^\s]+)", number_text)
                if match:
                    record["number"] = match.group(1)
        
        # Стадия
        stage_block = block.find("div", class_="registry-entry__header-mid__title")
        if stage_block:
            record["stage"] = stage_block.text.strip()
        
        # Название и организация
        body = block.find("div", class_="registry-entry__body")
        if body:
            name_block = body.find("div", class_="registry-entry__body-value")
            if name_block:
                name = name_block.text.strip()
                record["name"] = re.sub(r"\s+", " ", name)
            
            href_block = body.find("div", class_="registry-entry__body-href")
            if href_block:
                link_tag = href_block.find("a")
                if link_tag:
                    record["agency"] = link_tag.text.strip()
        
        # Цена и даты
        right_block = block.find("div", class_="registry-entry__right-block")
        if right_block:
            price_block = right_block.find("div", class_="price-block__value")
            if price_block:
                record["price"] = parse_price(price_block.text)
            
            data_block = right_block.find("div", class_="data-block")
            if data_block:
                for date_block in data_block.find_all("div", class_="data-block__item"):
                    title = date_block.find("div", class_="data-block__title")
                    value = date_block.find("div", class_="data-block__value")
                    if title and value:
                        title_text = title.text.strip()
                        date_val = parse_date(value.text)
                        if title_text == "Размещено":
                            record["published_date"] = date_val
                        elif title_text == "Обновлено":
                            record["updated_date"] = date_val
                        elif title_text == "Окончание подачи заявок":
                            record["end_date"] = date_val
        
        return record if record.get("number") else None
    except Exception as e:
        return None


def search_purchases(inn, date_from, date_to, page_size=50, max_pages=50, delay=0.5):
    """Поиск закупок по ИНН"""
    all_purchases = []
    page = 1
    found = set()
    
    print(f"Поиск для ИНН {inn}")
    
    while page <= max_pages:
        params = build_search_params(inn, date_from, date_to, page, page_size)
        
        try:
            response = requests.get(SEARCH_URL, headers=HEADERS, params=params, timeout=30)
            
            if response.status_code != 200:
                print(f"  Ошибка HTTP {response.status_code}")
                break
            
            soup = BeautifulSoup(response.text, "html.parser")
            
            if page == 1:
                total = soup.find("div", class_="search-results__total")
                if total:
                    print(f"  Найдено: {total.text.strip()}")
            
            blocks = soup.find_all("div", class_="search-registry-entry-block")
            if not blocks:
                break
            
            for block in blocks:
                p = extract_purchase_info(block)
                if p and p.get("number") not in found:
                    found.add(p.get("number"))
                    all_purchases.append(p)
            
            print(f"  Страница {page}: {len(blocks)} записей, добавлено {len(all_purchases)}")
            
            # Проверка следующей страницы
            pagination = soup.find("div", class_="search-results__pagination")
            if not pagination or not pagination.find("a", title="Следующая страница"):
                break
            
            page += 1
            time.sleep(delay)
            
        except Exception as e:
            print(f"  Ошибка: {e}")
            break
    
    print(f"  ИТОГО: {len(all_purchases)} закупок")
    return all_purchases


def save_to_excel(results, filename):
    """Сохранение в Excel"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    ws = wb.create_sheet(title="lots")
    
    headers = ["ID", "Номер закупки", "Название", "Организация", "Цена (руб.)",
               "Стадия", "ФЗ", "Тип", "Дата размещения", "Дата обновления",
               "Дата окончания", "Ссылка"]
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)
    
    row = 2
    for inn, purchases in results.items():
        for p in purchases:
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=p.get("number", ""))
            ws.cell(row=row, column=3, value=p.get("name", ""))
            ws.cell(row=row, column=4, value=p.get("agency", ""))
            ws.cell(row=row, column=5, value=p.get("price", 0))
            ws.cell(row=row, column=6, value=p.get("stage", ""))
            ws.cell(row=row, column=7, value=p.get("fz", ""))
            ws.cell(row=row, column=8, value=p.get("type", ""))
            
            if p.get("published_date"):
                ws.cell(row=row, column=9, value=p.get("published_date"))
            if p.get("updated_date"):
                ws.cell(row=row, column=10, value=p.get("updated_date"))
            if p.get("end_date"):
                ws.cell(row=row, column=11, value=p.get("end_date"))
            
            ws.cell(row=row, column=12, value=p.get("link", ""))
            row += 1
    
    # Автоширина
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
    
    wb.save(filename)
    return filename


def main():
    args = parse_arguments()
    
    # Определяем список ИНН
    if args.inn:
        # Если ИНН указаны в командной строке - используем их
        inn_list = args.inn
        inn_names = {inn: inn for inn in inn_list}
        print(f"Используем ИНН из командной строки: {len(inn_list)}")
    else:
        # Иначе загружаем из файла
        inn_names = load_inn_from_file()
        inn_list = list(inn_names.keys())
        print(f"Используем ИНН из inn_list.txt: {len(inn_list)}")
    
    if not inn_list:
        print("Ошибка: нет ИНН для поиска")
        print("Укажите ИНН через -i или добавьте в inn_list.txt")
        return
    
    # Определяем период
    date_from, date_to = get_date_range(args)
    
    print(f"\nПоиск закупок...")
    print("-" * 50)
    
    # Поиск по каждому ИНН
    results = {}
    for inn in inn_list:
        purchases = search_purchases(
            inn, date_from, date_to,
            args.page_size, args.max_pages, args.delay
        )
        results[inn] = purchases
    
    # Итоги
    print("\n" + "=" * 50)
    print("ИТОГИ:")
    total = 0
    for inn, purchases in results.items():
        name = inn_names.get(inn, inn)
        print(f"  {name}: {len(purchases)} закупок")
        total += len(purchases)
    print(f"  ВСЕГО: {total} закупок")
    print("=" * 50)
    
    # Сохранение
    if total > 0:
        filename = get_output_filename()
        save_to_excel(results, filename)
        print(f"\n✅ Результат сохранен: {filename}")
    else:
        print("\n❌ Закупок не найдено")

if __name__ == "__main__":
    main()

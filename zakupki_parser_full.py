#!/usr/bin/env python3
import sys
import os
import re
from datetime import datetime, date, timedelta
import argparse
import time
import logging
from pathlib import Path

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ZakupkiParser:
    def __init__(self):
        self.headers = {"user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        self.base_url = "https://zakupki.gov.ru"
        self.search_url = "https://zakupki.gov.ru/epz/order/extendedsearch/results.html"
        Path("output").mkdir(exist_ok=True)
    
    def load_inn_list(self):
        inn_dict = {}
        try:
            with open("inn_list.txt", 'r') as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#'):
                        parts = line.split('|')
                        if len(parts) >= 2:
                            inn_dict[parts[0]] = parts[1]
            return inn_dict
        except:
            return {}
    
    def get_output_filename(self):
        now = datetime.now()
        date_str = now.strftime("%d.%m.%Y")
        existing = [f for f in os.listdir("output") if f.startswith(f"zakupki_{date_str}_")]
        next_num = len(existing) + 1
        return f"output/zakupki_{date_str}_{next_num:02d}.xlsx"
    
    def parse_date(self, date_text):
        if not date_text:
            return None
        try:
            return datetime.strptime(date_text.strip(), "%d.%m.%Y").date()
        except:
            return None
    
    def parse_price(self, price_text):
        if not price_text:
            return 0.0
        cleaned = re.sub(r"[₽\s]", "", price_text)
        cleaned = cleaned.replace(",", ".").replace(chr(160), "")
        try:
            return float(cleaned)
        except:
            return 0.0
    
    def parse_purchase(self, block):
        try:
            record = {}
            
            # ФЗ и тип
            header = block.find("div", class_="registry-entry__header-top__title")
            if header:
                values = [e.strip() for e in header.text.split("\n") if e.strip()]
                if len(values) >= 2:
                    record["fz"] = values[0]
                    record["type"] = values[1]
            
            # Номер и ссылка
            number_div = block.find("div", class_="registry-entry__header-mid__number")
            if number_div:
                link = number_div.find("a")
                if link:
                    href = link.get("href", "")
                    record["link"] = self.base_url + href if record.get("fz") == "44-ФЗ" else href
                    text = link.text.strip()
                    match = re.search(r"№\s*([^\s]+)", text)
                    if match:
                        record["number"] = match.group(1)
            
            # Стадия
            stage_div = block.find("div", class_="registry-entry__header-mid__title")
            if stage_div:
                record["stage"] = stage_div.text.strip()
            
            # Название и организация
            body = block.find("div", class_="registry-entry__body")
            if body:
                name_div = body.find("div", class_="registry-entry__body-value")
                if name_div:
                    record["name"] = re.sub(r"\s+", " ", name_div.text.strip())
                
                agency_div = body.find("div", class_="registry-entry__body-href")
                if agency_div:
                    agency_link = agency_div.find("a")
                    if agency_link:
                        record["agency"] = agency_link.text.strip()
            
            # Цена и даты
            right = block.find("div", class_="registry-entry__right-block")
            if right:
                price_div = right.find("div", class_="price-block__value")
                if price_div:
                    record["price"] = self.parse_price(price_div.text)
                
                data_block = right.find("div", class_="data-block")
                if data_block:
                    for item in data_block.find_all("div", class_="data-block__item"):
                        title = item.find("div", class_="data-block__title")
                        value = item.find("div", class_="data-block__value")
                        if title and value:
                            title_text = title.text.strip()
                            date_val = self.parse_date(value.text)
                            if title_text == "Размещено":
                                record["published_date"] = date_val
                            elif title_text == "Обновлено":
                                record["updated_date"] = date_val
                            elif title_text == "Окончание подачи заявок":
                                record["end_date"] = date_val
            
            return record if "number" in record else None
        except Exception as e:
            return None
    
    def search_purchases(self, inn, date_from, date_to):
        all_purchases = []
        page = 1
        
        logger.info(f"Поиск для ИНН {inn}")
        
        while True:
            params = {
                "searchString": str(inn),
                "morphology": "on",
                "updateDateFrom": date_from.strftime("%d.%m.%Y"),
                "updateDateTo": date_to.strftime("%d.%m.%Y"),
                "sortBy": "ID",
                "sortDirection": "false",
                "recordsPerPage": "_50",
                "pageNumber": str(page),
                "fz44": "on",
                "fz223": "on",
                "af": "on",
                "ca": "on",
                "pc": "on",
                "pa": "on",
            }
            
            try:
                response = requests.get(self.search_url, headers=self.headers, params=params, timeout=30)
                if response.status_code != 200:
                    break
                
                soup = BeautifulSoup(response.text, "html.parser")
                blocks = soup.find_all("div", class_="search-registry-entry-block")
                
                if not blocks:
                    break
                
                for block in blocks:
                    purchase = self.parse_purchase(block)
                    if purchase:
                        all_purchases.append(purchase)
                
                pagination = soup.find("div", class_="search-results__pagination")
                if not pagination or not pagination.find("a", title="Следующая страница"):
                    break
                
                page += 1
                time.sleep(0.5)
                
            except Exception as e:
                logger.error(f"Ошибка: {e}")
                break
        
        logger.info(f"Найдено: {len(all_purchases)} закупок")
        return all_purchases
    
    def save_to_excel(self, results, filename, inn_names):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        columns = [
            "№", "Номер закупки", "Название", "Организация", "Цена (руб.)",
            "Стадия", "ФЗ", "Тип", "Дата размещения", "Дата обновления",
            "Дата окончания", "Ссылка"
        ]
        
        for inn, purchases in results.items():
            company_name = inn_names.get(inn, inn)
            sheet_name = company_name[:31]
            ws = wb.create_sheet(title=sheet_name)
            
            for col, header in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            for row, p in enumerate(purchases, 2):
                ws.cell(row=row, column=1, value=row-1)
                ws.cell(row=row, column=2, value=p.get("number", ""))
                ws.cell(row=row, column=3, value=p.get("name", ""))
                ws.cell(row=row, column=4, value=p.get("agency", ""))
                ws.cell(row=row, column=5, value=p.get("price", 0))
                ws.cell(row=row, column=6, value=p.get("stage", ""))
                ws.cell(row=row, column=7, value=p.get("fz", ""))
                ws.cell(row=row, column=8, value=p.get("type", ""))
                
                if p.get("published_date"):
                    ws.cell(row=row, column=9, value=p.get("published_date"))
                if p.get("updated_date"):
                    ws.cell(row=row, column=10, value=p.get("updated_date"))
                if p.get("end_date"):
                    ws.cell(row=row, column=11, value=p.get("end_date"))
                if p.get("link"):
                    ws.cell(row=row, column=12, value=p.get("link"))
            
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
        
        wb.save(filename)
        return filename
    
    def run(self, year=None, month=None):
        inn_dict = self.load_inn_list()
        inn_list = list(inn_dict.keys())
        
        now = datetime.now()
        if year and month:
            date_from = date(year, month, 1)
            if month == 12:
                date_to = date(year, 31, 12)
            else:
                date_to = date(year, month + 1, 1) - timedelta(days=1)
        else:
            date_from = date(now.year, 1, 1)
            date_to = date(now.year, 12, 31)
        
        logger.info(f"Период: {date_from} - {date_to}")
        
        results = {}
        for inn in inn_list:
            purchases = self.search_purchases(inn, date_from, date_to)
            results[inn] = purchases
            logger.info(f"ИНН {inn}: {len(purchases)} закупок")
        
        if any(results.values()):
            filename = self.get_output_filename()
            self.save_to_excel(results, filename, inn_dict)
            return results, filename
        return results, None

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("-y", "--year", type=int, help="Год")
    parser.add_argument("-m", "--month", type=int, help="Месяц")
    args = parser.parse_args()
    
    zakupki = ZakupkiParser()
    results, filename = zakupki.run(year=args.year, month=args.month)
    
    if filename:
        print(f"\n✅ Результат: {filename}")
        total = sum(len(r) for r in results.values())
        print(f"📊 Закупок: {total}")
    else:
        print("\n❌ Закупок не найдено")

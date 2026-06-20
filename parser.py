#!/usr/bin/env python3
"""
Парсер закупок для микросервиса - исправленная версия
"""

import sys
import os
import re
from datetime import datetime, date, timedelta
import argparse
import json
import time
import yaml
import glob

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Загрузка конфига
def load_config():
    with open('config.yaml', 'r') as f:
        return yaml.safe_load(f)

CONFIG = load_config()
HEADERS = {
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:97.0) Gecko/20100101 Firefox/97.0",
}
BASE_URL = "https://zakupki.gov.ru"
SEARCH_URL = "https://zakupki.gov.ru/epz/order/extendedsearch/results.html"

os.makedirs(CONFIG['excel']['output_dir'], exist_ok=True)


def load_inn_from_file():
    inn_dict = {}
    try:
        with open("inn_list.txt", 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    parts = line.split('|')
                    inn_dict[parts[0]] = parts[1] if len(parts) > 1 else parts[0]
        return inn_dict
    except FileNotFoundError:
        return {}


def parse_arguments():
    parser = argparse.ArgumentParser()
    parser.add_argument("-i", "--inn", nargs="+", help="ИНН (если не указаны - берем из inn_list.txt)")
    parser.add_argument("-y", "--year", type=int, help="Год")
    parser.add_argument("-m", "--month", type=int, help="Месяц")
    parser.add_argument("--date-from", help="Дата начала (ДД.ММ.ГГГГ)")
    parser.add_argument("--date-to", help="Дата окончания (ДД.ММ.ГГГГ)")
    parser.add_argument("--days", type=int, default=CONFIG['parser']['default_days'])
    parser.add_argument("--page-size", type=int, default=CONFIG['parser']['page_size'])
    parser.add_argument("--max-pages", type=int, default=CONFIG['parser']['max_pages'])
    parser.add_argument("--delay", type=float, default=CONFIG['parser']['delay'])
    parser.add_argument("--debug", action="store_true")
    return parser.parse_args()


def get_output_filename():
    """Генерирует имя файла в папке output/ДД.ММ.ГГГГ_NN/"""
    now = datetime.now()
    date_str = now.strftime("%d.%m.%Y")
    
    # Проверяем существующие папки за сегодня
    pattern = f"output/{date_str}_*"
    existing_folders = glob.glob(pattern)
    
    if not existing_folders:
        folder_number = 1
    else:
        numbers = []
        for folder in existing_folders:
            match = re.search(r'(\d{2}\.\d{2}\.\d{4})_(\d+)', folder)
            if match:
                numbers.append(int(match.group(2)))
        folder_number = max(numbers) + 1 if numbers else 1
    
    # Создаём папку
    folder_name = f"{date_str}_{folder_number:02d}"
    folder_path = os.path.join("output", folder_name)
    os.makedirs(folder_path, exist_ok=True)
    
    # Имя файла с номером внутри папки
    existing_files = [f for f in os.listdir(folder_path) if f.startswith(f"zakupki_{date_str}_")]
    file_number = len(existing_files) + 1
    
    filename = os.path.join(folder_path, f"zakupki_{date_str}_{file_number:02d}.xlsx")
    return filename

def get_date_range(args):
    now = datetime.now().date()
    
    if args.date_from and args.date_to:
        return (datetime.strptime(args.date_from, "%d.%m.%Y").date(),
                datetime.strptime(args.date_to, "%d.%m.%Y").date())
    
    if args.year:
        if args.month:
            date_from = date(args.year, args.month, 1)
            if args.month == 12:
                date_to = date(args.year, 12, 31)
            else:
                date_to = date(args.year, args.month + 1, 1) - timedelta(days=1)
        else:
            date_from = date(args.year, 1, 1)
            date_to = date(args.year, 12, 31)
        return date_from, date_to
    
    date_from = date(now.year, 1, 1)
    date_to = date(now.year, 12, 31)
    return date_from, date_to


def build_search_params(inn, date_from, date_to, page, page_size):
    return {
        "searchString": str(inn),
        "morphology": "on",
        "updateDateFrom": date_from.strftime("%d.%m.%Y"),
        "updateDateTo": date_to.strftime("%d.%m.%Y"),
        "sortBy": "ID",
        "sortDirection": "false",
        "recordsPerPage": f"_{page_size}",
        "pageNumber": str(page),
        "fz44": "on",
        "fz223": "on",
        "af": "on",
        "ca": "on",
        "pc": "on",
        "pa": "on",
        "showLotsInfoHidden": "false",
        "OrderPlacementSmallBusinessSubject": "on",
        "OrderPlacementRnpData": "on",
        "OrderPlacementExecutionRequirement": "on",
        "priceContractAdvantages44IdNameHidden": "%7B%7D",
        "priceContractAdvantages94IdNameHidden": "%7B%7D",
        "currencyIdGeneral": "-1",
        "selectedSubjectsIdNameHidden": "%7B%7D",
        "contractPriceCurrencyId": "-1",
        "budgetLevelIdNameHidden": "%7B%7D",
        "nonBudgetTypesIdNameHidden": "%7B%7D",
        "orderPlacement94_0": "0",
        "orderPlacement94_1": "0",
        "orderPlacement94_2": "0",
    }


def parse_price(price_text):
    if not price_text:
        return 0.0
    cleaned = re.sub(r"[₽\s]", "", price_text)
    cleaned = cleaned.replace(",", ".").replace(chr(160), "")
    try:
        return float(cleaned)
    except:
        return 0.0


def parse_date(date_text):
    if not date_text:
        return None
    try:
        return datetime.strptime(date_text.strip(), "%d.%m.%Y").date()
    except:
        return None


def extract_purchase_info(block):
    """Извлечение информации о закупке из HTML-блока"""
    try:
        record = {}
        
        # ФЗ и тип
        header = block.find("div", class_="registry-entry__header-top__title")
        if header:
            values = [e.strip() for e in header.text.split("\n") if e.strip()]
            if len(values) >= 2:
                record["fz"] = values[0]
                record["type"] = values[1]
        
        # Номер и ссылка
        number_block = block.find("div", class_="registry-entry__header-mid__number")
        if number_block:
            link_tag = number_block.find("a")
            if link_tag:
                href = link_tag.get("href", "")
                record["link"] = BASE_URL + href if record.get("fz") == "44-ФЗ" else href
                number_text = link_tag.text.strip()
                match = re.search(r"№\s*([^\s]+)", number_text)
                if match:
                    record["number"] = match.group(1)
        
        # Стадия
        stage_block = block.find("div", class_="registry-entry__header-mid__title")
        if stage_block:
            record["stage"] = stage_block.text.strip()
        
        # Название и организация
        body = block.find("div", class_="registry-entry__body")
        if body:
            name_block = body.find("div", class_="registry-entry__body-value")
            if name_block:
                record["name"] = re.sub(r"\s+", " ", name_block.text.strip())
            
            href_block = body.find("div", class_="registry-entry__body-href")
            if href_block:
                link_tag = href_block.find("a")
                if link_tag:
                    record["agency"] = link_tag.text.strip()
        
        # Цена и даты
        right_block = block.find("div", class_="registry-entry__right-block")
        if right_block:
            # Цена
            price_block = right_block.find("div", class_="price-block__value")
            if price_block:
                record["price"] = parse_price(price_block.text)
            
            # Даты - ищем data-block
            data_block = right_block.find("div", class_="data-block")
            if data_block:
                # Ищем Размещено и Обновлено (внутри row)
                row = data_block.find("div", class_="row")
                if row:
                    for col in row.find_all("div", class_="col-6"):
                        title = col.find("div", class_="data-block__title")
                        value = col.find("div", class_="data-block__value")
                        if title and value:
                            title_text = title.text.strip()
                            date_val = parse_date(value.text)
                            if title_text == "Размещено":
                                record["published_date"] = date_val
                            elif title_text == "Обновлено":
                                record["updated_date"] = date_val
                
                # Ищем Окончание подачи заявок (вне row)
                end_title = data_block.find("div", class_="data-block__title", string="Окончание подачи заявок")
                if end_title:
                    end_value = end_title.find_next_sibling("div", class_="data-block__value")
                    if end_value:
                        record["end_date"] = parse_date(end_value.text)
        
        return record if record.get("number") else None
    except Exception as e:
        if CONFIG['parser']['debug']:
            print(f"Ошибка парсинга: {e}")
        return None


def search_purchases(inn, date_from, date_to, page_size, max_pages, delay):
    all_purchases = []
    page = 1
    found_numbers = set()
    
    print(f"\nПоиск для ИНН {inn}")
    
    while page <= max_pages:
        params = build_search_params(inn, date_from, date_to, page, page_size)
        
        try:
            response = requests.get(SEARCH_URL, headers=HEADERS, params=params, timeout=30)
            
            if response.status_code != 200:
                print(f"  Ошибка HTTP {response.status_code}")
                break
            
            soup = BeautifulSoup(response.text, "html.parser")
            
            if page == 1:
                total_block = soup.find("div", class_="search-results__total")
                if total_block:
                    print(f"  Найдено: {total_block.text.strip()}")
            
            blocks = soup.find_all("div", class_="search-registry-entry-block")
            if not blocks:
                print(f"  Записей не найдено")
                break
            
            page_purchases = []
            for block in blocks:
                purchase = extract_purchase_info(block)
                if purchase and purchase.get("number") not in found_numbers:
                    found_numbers.add(purchase.get("number"))
                    page_purchases.append(purchase)
            
            all_purchases.extend(page_purchases)
            print(f"  Страница {page}: {len(page_purchases)} записей (всего {len(all_purchases)})")
            
            pagination = soup.find("div", class_="search-results__pagination")
            next_link = None
            if pagination:
                next_link = pagination.find("a", title="Следующая страница")
            
            if not next_link:
                break
            
            page += 1
            
            if delay > 0:
                time.sleep(delay)
                
        except Exception as e:
            print(f"  Ошибка: {e}")
            break
    
    return all_purchases


def save_to_excel(results, filename, script_start_time, script_name, inn_names):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Лист lots
    ws_lots = wb.create_sheet(title="lots")
    
    headers = [
        "ID", "Номер закупки", "Название", "Организация", "Цена (руб.)",
        "Стадия", "ФЗ", "Тип", "Дата размещения", "Дата обновления",
        "Дата окончания", "Ссылка"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws_lots.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    row = 2
    for inn, purchases in results.items():
        for purchase in purchases:
            ws_lots.cell(row=row, column=1, value=row-1)
            ws_lots.cell(row=row, column=2, value=purchase.get("number", ""))
            ws_lots.cell(row=row, column=3, value=purchase.get("name", ""))
            ws_lots.cell(row=row, column=4, value=purchase.get("agency", ""))
            
            price_cell = ws_lots.cell(row=row, column=5, value=purchase.get("price", 0))
            price_cell.number_format = '#,##0.00₽'
            
            ws_lots.cell(row=row, column=6, value=purchase.get("stage", ""))
            ws_lots.cell(row=row, column=7, value=purchase.get("fz", ""))
            ws_lots.cell(row=row, column=8, value=purchase.get("type", ""))
            
            if purchase.get("published_date"):
                ws_lots.cell(row=row, column=9, value=purchase.get("published_date"))
            if purchase.get("updated_date"):
                ws_lots.cell(row=row, column=10, value=purchase.get("updated_date"))
            if purchase.get("end_date"):
                ws_lots.cell(row=row, column=11, value=purchase.get("end_date"))
            
            ws_lots.cell(row=row, column=12, value=purchase.get("link", ""))
            row += 1
    
    # Автоширина
    for col in ws_lots.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws_lots.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # Лист log
    ws_log = wb.create_sheet(title="log")
    
    log_headers = ["Date time", "Script name", "Script mod time", "Command line", "Stage", "INN list", "Years"]
    for col, header in enumerate(log_headers, 1):
        cell = ws_log.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    script_mod_time = datetime.fromtimestamp(os.path.getmtime(__file__))
    inn_list = list(results.keys())
    
    ws_log.cell(row=2, column=1, value=script_start_time)
    ws_log.cell(row=2, column=2, value=script_name)
    ws_log.cell(row=2, column=3, value=script_mod_time)
    ws_log.cell(row=2, column=4, value=json.dumps(sys.argv))
    ws_log.cell(row=2, column=5, value=1)
    ws_log.cell(row=2, column=6, value=json.dumps(inn_list))
    ws_log.cell(row=2, column=7, value="[]")
    
    for col in ws_log.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws_log.column_dimensions[col_letter].width = min(max_length + 2, 30)
    
    wb.save(filename)
    return True


def main():
    script_start_time = datetime.now()
    script_name = os.path.basename(__file__)
    
    print("=" * 60)
    print("ПАРСЕР ЗАКУПОК (микросервис)")
    print(f"Время запуска: {script_start_time.strftime('%d.%m.%Y %H:%M:%S')}")
    print("=" * 60)
    
    args = parse_arguments()
    date_from, date_to = get_date_range(args)
    
    if args.inn:
        inn_dict = {inn: inn for inn in args.inn}
        print(f"ИНН из командной строки: {len(inn_dict)}")
    else:
        inn_dict = load_inn_from_file()
        if not inn_dict:
            print("Ошибка: нет ИНН для поиска")
            return
        print(f"ИНН из inn_list.txt: {len(inn_dict)}")
    
    print(f"Период: {date_from.strftime('%d.%m.%Y')} - {date_to.strftime('%d.%m.%Y')}")
    print(f"Страниц максимум: {args.max_pages}")
    print(f"Записей на страницу: {args.page_size}")
    
    results = {}
    for inn, name in inn_dict.items():
        purchases = search_purchases(
            inn, date_from, date_to,
            args.page_size, args.max_pages, args.delay
        )
        results[inn] = purchases
        print(f"  ИНН {inn}: {len(purchases)} закупок")
    
    total = sum(len(p) for p in results.values())
    if total > 0:
        filename = get_output_filename()
        save_to_excel(results, filename, script_start_time, script_name, inn_dict)
        print(f"\n✅ Результат: {filename}")
        print(f"📊 Всего закупок: {total}")
    else:
        print("\n❌ Закупок не найдено")

if __name__ == "__main__":
    main()

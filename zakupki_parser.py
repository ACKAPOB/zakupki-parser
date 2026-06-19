#!/usr/bin/env python3
import sys
import os
import re
from datetime import datetime, date, timedelta
import argparse
import time
import json
import yaml
import logging
from pathlib import Path

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ZakupkiParser:
    def __init__(self, config_path="config.yaml"):
        with open(config_path, 'r') as f:
            self.config = yaml.safe_load(f)
        
        self.headers = {
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        }
        self.base_url = "https://zakupki.gov.ru"
        self.search_url = "https://zakupki.gov.ru/epz/order/extendedsearch/results.html"
        
        # Создаем директории
        Path("output").mkdir(exist_ok=True)
        Path("logs").mkdir(exist_ok=True)
    
    def load_inn_list(self):
        """Загрузка списка ИНН из файла"""
        inn_dict = {}
        try:
            with open("inn_list.txt", 'r') as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#'):
                        parts = line.split('|')
                        if len(parts) >= 2:
                            inn_dict[parts[0]] = parts[1]
            logger.info(f"Загружено {len(inn_dict)} ИНН")
            return inn_dict
        except Exception as e:
            logger.error(f"Ошибка загрузки ИНН: {e}")
            return {}
    
    def get_output_filename(self):
        """Генерация имени файла"""
        now = datetime.now()
        date_str = now.strftime("%d.%m.%Y")
        existing = [f for f in os.listdir("output") if f.startswith(f"zakupki_{date_str}_")]
        next_num = len(existing) + 1
        return f"output/zakupki_{date_str}_{next_num:02d}.xlsx"
    
    def search_purchases(self, inn, date_from, date_to):
        """Поиск закупок"""
        all_purchases = []
        page = 1
        max_pages = self.config['parser']['max_pages']
        
        logger.info(f"Поиск для ИНН {inn} за период {date_from} - {date_to}")
        
        while page <= max_pages:
            params = {
                "searchString": str(inn),
                "morphology": "on",
                "updateDateFrom": date_from.strftime("%d.%m.%Y"),
                "updateDateTo": date_to.strftime("%d.%m.%Y"),
                "sortBy": "ID",
                "sortDirection": "false",
                "recordsPerPage": f"_{self.config['parser']['page_size']}",
                "pageNumber": str(page),
                "fz44": "on",
                "fz223": "on",
                "af": "on",
                "ca": "on",
                "pc": "on",
                "pa": "on",
            }
            
            try:
                response = requests.get(self.search_url, headers=self.headers, params=params, timeout=30)
                
                if response.status_code != 200:
                    logger.error(f"HTTP {response.status_code}")
                    break
                
                soup = BeautifulSoup(response.text, "html.parser")
                blocks = soup.find_all("div", class_="search-registry-entry-block")
                
                if not blocks:
                    break
                
                for block in blocks:
                    purchase = self.parse_block(block)
                    if purchase:
                        all_purchases.append(purchase)
                
                logger.info(f"Страница {page}: {len(blocks)} записей")
                
                # Проверка на следующую страницу
                pagination = soup.find("div", class_="search-results__pagination")
                if not pagination or "Следующая" not in str(pagination):
                    break
                
                page += 1
                time.sleep(self.config['parser']['delay'])
                
            except Exception as e:
                logger.error(f"Ошибка: {e}")
                break
        
        return all_purchases
    
    def parse_block(self, block):
        """Парсинг одного блока закупки"""
        try:
            record = {}
            
            # Номер
            number_div = block.find("div", class_="registry-entry__header-mid__number")
            if number_div:
                link = number_div.find("a")
                if link:
                    text = link.text.strip()
                    match = re.search(r"№\s*([^\s]+)", text)
                    if match:
                        record["number"] = match.group(1)
            
            # Название
            body = block.find("div", class_="registry-entry__body")
            if body:
                name_div = body.find("div", class_="registry-entry__body-value")
                if name_div:
                    record["name"] = name_div.text.strip()
            
            # Цена
            right = block.find("div", class_="registry-entry__right-block")
            if right:
                price_div = right.find("div", class_="price-block__value")
                if price_div:
                    price_text = price_div.text.strip()
                    price_text = re.sub(r"[₽\s]", "", price_text).replace(",", ".")
                    try:
                        record["price"] = float(price_text)
                    except:
                        record["price"] = 0
            
            return record if "number" in record else None
            
        except Exception as e:
            return None
    
    def save_to_excel(self, results, filename, inn_names):
        """Сохранение в Excel"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        for inn, purchases in results.items():
            name = inn_names.get(inn, inn)
            ws = wb.create_sheet(title=name[:31])
            
            headers = ["№", "Номер закупки", "Название", "Цена"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            for row, p in enumerate(purchases, 2):
                ws.cell(row=row, column=1, value=row-1)
                ws.cell(row=row, column=2, value=p.get("number", ""))
                ws.cell(row=row, column=3, value=p.get("name", ""))
                ws.cell(row=row, column=4, value=p.get("price", 0))
        
        wb.save(filename)
        logger.info(f"Сохранено: {filename}")
        return filename
    
    def run(self, inn_list=None, year=None, month=None):
        """Запуск парсера"""
        # Загружаем ИНН
        if not inn_list:
            inn_dict = self.load_inn_list()
            inn_list = list(inn_dict.keys())
        else:
            inn_dict = {inn: inn for inn in inn_list}
        
        # Определяем период
        now = datetime.now()
        if year and month:
            date_from = date(year, month, 1)
            if month == 12:
                date_to = date(year, 31, 12)
            else:
                date_to = date(year, month+1, 1) - timedelta(days=1)
        else:
            date_from = date(now.year, 1, 1)
            date_to = date(now.year, 12, 31)
        
        # Поиск
        results = {}
        for inn in inn_list:
            purchases = self.search_purchases(inn, date_from, date_to)
            results[inn] = purchases
            logger.info(f"ИНН {inn}: {len(purchases)} закупок")
        
        # Сохранение
        if any(results.values()):
            filename = self.get_output_filename()
            self.save_to_excel(results, filename, inn_dict)
            return results, filename
        
        return results, None

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("-i", "--inn", nargs="+", help="Список ИНН")
    parser.add_argument("-y", "--year", type=int, help="Год")
    parser.add_argument("-m", "--month", type=int, help="Месяц")
    args = parser.parse_args()
    
    zakupki = ZakupkiParser()
    results, filename = zakupki.run(inn_list=args.inn, year=args.year, month=args.month)
    
    if filename:
        print(f"\n✅ Результат: {filename}")
        print(f"📊 Закупок: {sum(len(r) for r in results.values())}")
    else:
        print("\n❌ Закупок не найдено")

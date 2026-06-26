#!/usr/bin/env python3
"""
Расширенный парсер закупок с детальной информацией о лотах
Дополняет данные из parser.py информацией с детальных страниц лотов
"""

import sys
import os
import re
from datetime import datetime, date, timedelta
import argparse
import json
import time
import yaml
import glob
import logging

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ===== НАСТРОЙКА ЛОГИРОВАНИЯ =====
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('parser_extended.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


# ===== ЗАГРУЗКА КОНФИГА =====
def load_config():
    try:
        with open('config.yaml', 'r', encoding='utf-8') as f:
            return yaml.safe_load(f)
    except Exception as e:
        logger.error(f"Ошибка загрузки конфига: {e}")
        return {}

CONFIG = load_config()

# ===== КОНСТАНТЫ =====
HEADERS = {
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:97.0) Gecko/20100101 Firefox/97.0",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "ru-RU,ru;q=0.8,en-US;q=0.5,en;q=0.3",
}
BASE_URL = "https://zakupki.gov.ru"
SEARCH_URL = "https://zakupki.gov.ru/epz/order/extendedsearch/results.html"

# Настройки из конфига
ext_config = CONFIG.get('parser_extended', {})
DELAY_BETWEEN_INN = ext_config.get('delay_between_inn', 3)
DELAY_BETWEEN_PURCHASES = ext_config.get('delay_between_purchases', 1)
DELAY_BETWEEN_LOTS = ext_config.get('delay_between_lots', 1)
DELAY_RETRY = ext_config.get('delay_retry', 5)
MAX_RETRIES = ext_config.get('max_retries', 3)
REQUEST_TIMEOUT = ext_config.get('request_timeout', 30)

# Создаём папку для результатов (из общего конфига)
OUTPUT_DIR = CONFIG.get('excel', {}).get('output_dir', 'output')
os.makedirs(OUTPUT_DIR, exist_ok=True)


# ===== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ =====

def load_inn_from_file():
    """Загрузка списка ИНН из файла"""
    inn_dict = {}
    try:
        with open("inn_list.txt", 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    parts = line.split('|')
                    inn_dict[parts[0]] = parts[1] if len(parts) > 1 else parts[0]
        return inn_dict
    except FileNotFoundError:
        logger.error("Файл inn_list.txt не найден")
        return {}


def parse_arguments():
    parser = argparse.ArgumentParser(description="Расширенный парсер закупок")
    parser.add_argument("-i", "--inn", nargs="+", help="ИНН (если не указаны - из inn_list.txt)")
    parser.add_argument("-y", "--year", type=int, help="Год")
    parser.add_argument("-m", "--month", type=int, help="Месяц")
    parser.add_argument("--date-from", help="Дата начала (ДД.ММ.ГГГГ)")
    parser.add_argument("--date-to", help="Дата окончания (ДД.ММ.ГГГГ)")
    parser.add_argument("--days", type=int, default=CONFIG.get('parser', {}).get('default_days', 365))
    parser.add_argument("--page-size", type=int, default=CONFIG.get('parser', {}).get('page_size', 50))
    parser.add_argument("--max-pages", type=int, default=CONFIG.get('parser', {}).get('max_pages', 10))
    parser.add_argument("--delay", type=float, default=CONFIG.get('parser', {}).get('delay', 0.5))
    parser.add_argument("--skip-lots", action="store_true", help="Не собирать детали лотов")
    parser.add_argument("--debug", action="store_true")
    return parser.parse_args()


def get_output_folder():
    """Создаёт папку результата. Если задана ZAKUPKI_OUTPUT_DIR - пишет туда."""
    # Проверяем переменную окружения от run_all.py
    session_dir = os.environ.get('ZAKUPKI_OUTPUT_DIR')
    
    if session_dir and os.path.isdir(session_dir):
        folder_path = session_dir
        logger.info(f"📁 Используем папку сессии: {folder_path}")
        return folder_path, 0  # Номер не важен
    
    # Обычная логика для отдельного запуска
    now = datetime.now()
    date_str = now.strftime("%d.%m.%Y")
    
    pattern = f"{OUTPUT_DIR}/{date_str}_*"
    existing_folders = glob.glob(pattern)
    
    if not existing_folders:
        folder_number = 1
    else:
        numbers = []
        for folder in existing_folders:
            match = re.search(r'(\d{2}\.\d{2}\.\d{4})_(\d+)', folder)
            if match:
                numbers.append(int(match.group(2)))
        folder_number = max(numbers) + 1 if numbers else 1
    
    folder_name = f"{date_str}_{folder_number:02d}"
    folder_path = os.path.join(OUTPUT_DIR, folder_name)
    os.makedirs(folder_path, exist_ok=True)
    
    logger.info(f"📁 Папка результата: {folder_path}")
    return folder_path, folder_number


def get_date_range(args):
    """Определение диапазона дат"""
    now = datetime.now().date()
    
    if args.date_from and args.date_to:
        return (datetime.strptime(args.date_from, "%d.%m.%Y").date(),
                datetime.strptime(args.date_to, "%d.%m.%Y").date())
    
    if args.year:
        if args.month:
            date_from = date(args.year, args.month, 1)
            if args.month == 12:
                date_to = date(args.year, 12, 31)
            else:
                date_to = date(args.year, args.month + 1, 1) - timedelta(days=1)
        else:
            date_from = date(args.year, 1, 1)
            date_to = date(args.year, 12, 31)
        return date_from, date_to
    
    date_from = date(now.year, 1, 1)
    date_to = date(now.year, 12, 31)
    return date_from, date_to


def parse_price(price_text):
    """Парсинг цены из текста"""
    if not price_text:
        return 0.0
    cleaned = re.sub(r"[₽\s]", "", price_text)
    cleaned = cleaned.replace(",", ".").replace(chr(160), "")
    try:
        return float(cleaned)
    except:
        return 0.0


def parse_date(date_text):
    """Парсинг даты из текста"""
    if not date_text:
        return None
    try:
        return datetime.strptime(date_text.strip(), "%d.%m.%Y").date()
    except:
        return None


def safe_request(url, params=None, retries=MAX_RETRIES):
    """Безопасный HTTP-запрос с повторными попытками"""
    for attempt in range(1, retries + 1):
        try:
            if attempt > 1:
                logger.info(f"      🔄 Попытка {attempt}/{retries}...")
                time.sleep(DELAY_RETRY)
            
            response = requests.get(
                url, 
                headers=HEADERS, 
                params=params, 
                timeout=REQUEST_TIMEOUT
            )
            
            if response.status_code == 200:
                return response
            else:
                logger.warning(f"      ⚠️  HTTP {response.status_code}")
                
        except Exception as e:
            logger.warning(f"      ⚠️  Ошибка: {e}")
    
    return None


# ===== ПАРСИНГ ЗАКУПОК (как в parser.py) =====

def build_search_params(inn, date_from, date_to, page, page_size):
    return {
        "searchString": str(inn),
        "morphology": "on",
        "updateDateFrom": date_from.strftime("%d.%m.%Y"),
        "updateDateTo": date_to.strftime("%d.%m.%Y"),
        "sortBy": "ID",
        "sortDirection": "false",
        "recordsPerPage": f"_{page_size}",
        "pageNumber": str(page),
        "fz44": "on",
        "fz223": "on",
        "af": "on",
        "ca": "on",
        "pc": "on",
        "pa": "on",
        "showLotsInfoHidden": "false",
        "OrderPlacementSmallBusinessSubject": "on",
        "OrderPlacementRnpData": "on",
        "OrderPlacementExecutionRequirement": "on",
        "priceContractAdvantages44IdNameHidden": "%7B%7D",
        "priceContractAdvantages94IdNameHidden": "%7B%7D",
        "currencyIdGeneral": "-1",
        "selectedSubjectsIdNameHidden": "%7B%7D",
        "contractPriceCurrencyId": "-1",
        "budgetLevelIdNameHidden": "%7B%7D",
        "nonBudgetTypesIdNameHidden": "%7B%7D",
        "orderPlacement94_0": "0",
        "orderPlacement94_1": "0",
        "orderPlacement94_2": "0",
    }


def extract_purchase_info(block):
    """Извлечение информации о закупке из HTML-блока"""
    try:
        record = {}
        
        # ФЗ и тип
        header = block.find("div", class_="registry-entry__header-top__title")
        if header:
            values = [e.strip() for e in header.text.split("\n") if e.strip()]
            if len(values) >= 2:
                record["fz"] = values[0]
                record["type"] = values[1]
        
        # Номер и ссылка
        number_block = block.find("div", class_="registry-entry__header-mid__number")
        if number_block:
            link_tag = number_block.find("a")
            if link_tag:
                href = link_tag.get("href", "")
                record["link"] = BASE_URL + href if not href.startswith('http') else href
                number_text = link_tag.text.strip()
                match = re.search(r"№\s*([^\s]+)", number_text)
                if match:
                    record["number"] = match.group(1)
                
                # Пытаемся извлечь noticeGuid из ссылки
                guid_match = re.search(r'noticeId=(\d+)', href)
                if guid_match:
                    record["notice_id"] = guid_match.group(1)
        
        # Стадия
        stage_block = block.find("div", class_="registry-entry__header-mid__title")
        if stage_block:
            record["stage"] = stage_block.text.strip()
        
        # Название и организация
        body = block.find("div", class_="registry-entry__body")
        if body:
            name_block = body.find("div", class_="registry-entry__body-value")
            if name_block:
                record["name"] = re.sub(r"\s+", " ", name_block.text.strip())
            
            href_block = body.find("div", class_="registry-entry__body-href")
            if href_block:
                link_tag = href_block.find("a")
                if link_tag:
                    record["agency"] = link_tag.text.strip()
        
        # Цена и даты
        right_block = block.find("div", class_="registry-entry__right-block")
        if right_block:
            price_block = right_block.find("div", class_="price-block__value")
            if price_block:
                record["price"] = parse_price(price_block.text)
            
            data_block = right_block.find("div", class_="data-block")
            if data_block:
                row = data_block.find("div", class_="row")
                if row:
                    for col in row.find_all("div", class_="col-6"):
                        title = col.find("div", class_="data-block__title")
                        value = col.find("div", class_="data-block__value")
                        if title and value:
                            title_text = title.text.strip()
                            date_val = parse_date(value.text)
                            if title_text == "Размещено":
                                record["published_date"] = date_val
                            elif title_text == "Обновлено":
                                record["updated_date"] = date_val
                
                end_title = data_block.find("div", class_="data-block__title", string="Окончание подачи заявок")
                if end_title:
                    end_value = end_title.find_next_sibling("div", class_="data-block__value")
                    if end_value:
                        record["end_date"] = parse_date(end_value.text)
        
        return record if record.get("number") else None
    except Exception as e:
        logger.warning(f"Ошибка парсинга закупки: {e}")
        return None


def search_purchases(inn, date_from, date_to, page_size, max_pages, delay):
    """Поиск закупок по ИНН"""
    all_purchases = []
    page = 1
    found_numbers = set()
    
    logger.info(f"\n🔍 Поиск закупок для ИНН {inn}")
    
    while page <= max_pages:
        params = build_search_params(inn, date_from, date_to, page, page_size)
        
        try:
            response = safe_request(SEARCH_URL, params)
            if not response:
                break
            
            soup = BeautifulSoup(response.text, "html.parser")
            
            if page == 1:
                total_block = soup.find("div", class_="search-results__total")
                if total_block:
                    logger.info(f"   Найдено: {total_block.text.strip()}")
            
            blocks = soup.find_all("div", class_="search-registry-entry-block")
            if not blocks:
                logger.info(f"   Записей не найдено")
                break
            
            page_purchases = []
            for block in blocks:
                purchase = extract_purchase_info(block)
                if purchase and purchase.get("number") not in found_numbers:
                    found_numbers.add(purchase.get("number"))
                    page_purchases.append(purchase)
            
            all_purchases.extend(page_purchases)
            logger.info(f"   Страница {page}: {len(page_purchases)} записей (всего {len(all_purchases)})")
            
            pagination = soup.find("div", class_="search-results__pagination")
            next_link = None
            if pagination:
                next_link = pagination.find("a", title="Следующая страница")
            
            if not next_link:
                break
            
            page += 1
            
            if delay > 0:
                time.sleep(delay)
                
        except Exception as e:
            logger.error(f"   ❌ Ошибка: {e}")
            break
    
    return all_purchases


# ===== ПАРСИНГ ДЕТАЛЕЙ ЛОТОВ =====

def get_notice_guid_from_purchase(purchase):
    """Получить noticeGuid со страницы закупки"""
    link = purchase.get("link", "")
    if not link:
        return None
    
    # Пытаемся извлечь из URL
    guid_match = re.search(r'noticeId=(\d+)', link)
    if guid_match:
        return guid_match.group(1)
    
    # Если не получилось - загружаем страницу
    try:
        response = safe_request(link)
        if not response:
            return None
        
        soup = BeautifulSoup(response.text, "html.parser")
        
        # Ищем noticeGuid в JavaScript или meta-тегах
        page_text = soup.prettify()
        guid_match = re.search(r'noticeGuid["\s:=]+["\']?([a-f0-9-]+)["\']?', page_text, re.I)
        if guid_match:
            return guid_match.group(1)
        
        # Ищем в ссылках на лоты
        lot_link = soup.find('a', href=re.compile(r'lot-list\.html.*noticeGuid='))
        if lot_link:
            href = lot_link.get('href', '')
            guid_match = re.search(r'noticeGuid=([a-f0-9-]+)', href, re.I)
            if guid_match:
                return guid_match.group(1)
        
    except Exception as e:
        logger.warning(f"      ⚠️  Ошибка получения noticeGuid: {e}")
    
    return None


def get_lot_list_url(purchase_number, fz_type):
    """Формирует URL списка лотов для закупки"""
    # Определяем тип извещения по ФЗ
    if "44" in fz_type:
        notice_type = "noticeCommonInfo"  # Для 44-ФЗ
    else:
        notice_type = "noticeCommonInfo"  # Для 223-ФЗ
    
    # Базовый URL для просмотра извещения
    return f"{BASE_URL}/epz/order/notice/{notice_type}.html?regNumber={purchase_number}"


def get_lot_info_url(lot_guid, notice_guid, purchase_number, fz_type):
    """Формирует URL сведений о лоте"""
    # Для 223-ФЗ
    if "223" in fz_type:
        return f"{BASE_URL}/epz/order/notice/notice223/lot/lot-info.html?lotGuid={lot_guid}&purchaseNoticeGuid={notice_guid}&purchaseNoticeNumber={purchase_number}"
    # Для 44-ФЗ
    elif "44" in fz_type:
        return f"{BASE_URL}/epz/order/notice/notice44/lot/lot-info.html?lotGuid={lot_guid}&purchaseNoticeGuid={notice_guid}&purchaseNoticeNumber={purchase_number}"
    return None


def parse_lot_list(html, purchase_number, fz_type):
    """Парсит список лотов со страницы закупки"""
    lots = []
    soup = BeautifulSoup(html, "html.parser")
    
    # Ищем таблицу с лотами или блоки с лотами
    # Вариант 1: таблица
    tables = soup.find_all('table')
    for table in tables:
        rows = table.find_all('tr')
        for row in rows:
            cells = row.find_all('td')
            if len(cells) >= 2:
                # Ищем ссылку на лот
                link = row.find('a', href=re.compile(r'lotGuid='))
                if link:
                    href = link.get('href', '')
                    lot_guid_match = re.search(r'lotGuid=([a-f0-9-]+)', href, re.I)
                    if lot_guid_match:
                        lots.append({
                            'lot_guid': lot_guid_match.group(1),
                            'lot_number': cells[0].text.strip() if cells else '',
                            'lot_name': link.text.strip()
                        })
    
    # Вариант 2: блоки с лотами
    lot_blocks = soup.find_all('div', class_=re.compile(r'lot|block', re.I))
    for block in lot_blocks:
        link = block.find('a', href=re.compile(r'lotGuid='))
        if link:
            href = link.get('href', '')
            lot_guid_match = re.search(r'lotGuid=([a-f0-9-]+)', href, re.I)
            if lot_guid_match:
                lots.append({
                    'lot_guid': lot_guid_match.group(1),
                    'lot_number': '',
                    'lot_name': link.text.strip()
                })
    
    # Вариант 3: ищем все ссылки на lot-info.html
    if not lots:
        links = soup.find_all('a', href=re.compile(r'lot-info\.html'))
        for link in links:
            href = link.get('href', '')
            lot_guid_match = re.search(r'lotGuid=([a-f0-9-]+)', href, re.I)
            if lot_guid_match:
                lots.append({
                    'lot_guid': lot_guid_match.group(1),
                    'lot_number': '',
                    'lot_name': link.text.strip()
                })
    
    return lots


def parse_lot_info(html):
    """Парсит детальную информацию о лоте"""
    info = {}
    soup = BeautifulSoup(html, "html.parser")
    
    # Ищем все блоки с информацией
    # Обычно это div с классом "blockInfo" или "registry-entry__body"
    
    # Вариант 1: ищем по заголовкам
    all_titles = soup.find_all(['div', 'span', 'td'], class_=re.compile(r'title|label|header', re.I))
    
    for title in all_titles:
        title_text = title.text.strip().lower()
        
        # Ищем значение - обычно следующий sibling или следующий элемент
        value = None
        next_elem = title.find_next_sibling()
        if next_elem:
            value = next_elem.text.strip()
        
        if not value:
            parent = title.parent
            if parent:
                value_elem = parent.find(class_=re.compile(r'value', re.I))
                if value_elem:
                    value = value_elem.text.strip()
        
        if not value:
            continue
        
        # Сопоставляем с нужными полями
        if 'номер плана' in title_text and 'позиции' not in title_text:
            info['plan_number'] = value
        elif 'номер позиции плана' in title_text:
            info['plan_position_number'] = value
        elif 'начальная' in title_text and 'цена' in title_text:
            info['initial_price'] = parse_price(value)
        elif 'период планирования' in title_text:
            info['planning_period'] = value
        elif 'размещено' in title_text and 'обновлено' not in title_text:
            info['published_date'] = parse_date(value)
        elif 'обновлено' in title_text:
            info['updated_date'] = parse_date(value)
        elif 'номер лота' in title_text:
            info['lot_number'] = value
        elif 'наименование предмета' in title_text or 'предмет договора' in title_text:
            info['lot_name'] = value
        elif 'место поставки' in title_text and 'адрес' not in title_text:
            info['delivery_region'] = value
        elif 'место поставки' in title_text and 'адрес' in title_text:
            info['delivery_address'] = value
        elif 'дата начала планирования' in title_text:
            info['planning_start_date'] = parse_date(value)
        elif 'дата окончания планирования' in title_text:
            info['planning_end_date'] = parse_date(value)
        elif 'позиция плана изменена' in title_text:
            info['plan_position_changed'] = value
    
    # Вариант 2: ищем по структуре data-block
    data_blocks = soup.find_all('div', class_='data-block__item')
    for block in data_blocks:
        title = block.find('div', class_='data-block__title')
        value = block.find('div', class_='data-block__value')
        if title and value:
            title_text = title.text.strip().lower()
            value_text = value.text.strip()
            
            if 'номер плана' in title_text and 'позиции' not in title_text:
                info['plan_number'] = value_text
            elif 'номер позиции плана' in title_text:
                info['plan_position_number'] = value_text
            elif 'начальная' in title_text and 'цена' in title_text:
                info['initial_price'] = parse_price(value_text)
            elif 'период планирования' in title_text:
                info['planning_period'] = value_text
            elif 'размещено' in title_text:
                info['published_date'] = parse_date(value_text)
            elif 'обновлено' in title_text:
                info['updated_date'] = parse_date(value_text)
            elif 'номер лота' in title_text:
                info['lot_number'] = value_text
            elif 'наименование предмета' in title_text:
                info['lot_name'] = value_text
            elif 'место поставки' in title_text and 'субъект' in title_text.lower():
                info['delivery_region'] = value_text
            elif 'место поставки' in title_text and 'адрес' in title_text.lower():
                info['delivery_address'] = value_text
    
    # Вариант 3: парсим через регулярки весь текст страницы
    page_text = soup.get_text()
    
    # Номер плана
    plan_match = re.search(r'Номер плана[:\s]+(\d+)', page_text)
    if plan_match and 'plan_number' not in info:
        info['plan_number'] = plan_match.group(1)
    
    # Номер позиции плана
    pos_match = re.search(r'Номер позиции плана[:\s]+(\d+)', page_text)
    if pos_match and 'plan_position_number' not in info:
        info['plan_position_number'] = pos_match.group(1)
    
    # Начальная цена
    price_match = re.search(r'Начальная.*?цена.*?(\d[\d\s,\.]+)₽', page_text, re.I | re.S)
    if price_match and 'initial_price' not in info:
        info['initial_price'] = parse_price(price_match.group(1))
    
    # Период планирования
    period_match = re.search(r'Период планирования[:\s]+([^\n]+)', page_text)
    if period_match and 'planning_period' not in info:
        info['planning_period'] = period_match.group(1).strip()
    
    # Место поставки (субъект РФ)
    region_match = re.search(r'Место поставки.*?субъект.*?РФ[:\s]+([^\n]+)', page_text, re.I | re.S)
    if region_match and 'delivery_region' not in info:
        info['delivery_region'] = region_match.group(1).strip()
    
    # Место поставки (адрес)
    address_match = re.search(r'Место поставки.*?адрес[:\s]+([^\n]+)', page_text, re.I | re.S)
    if address_match and 'delivery_address' not in info:
        info['delivery_address'] = address_match.group(1).strip()
    
    return info


def fetch_purchase_lot_details(purchase, skip_lots=False):
    """Получить детальную информацию по всем лотам закупки"""
    if skip_lots:
        return []
    
    purchase_number = purchase.get("number", "")
    fz_type = purchase.get("fz", "")
    link = purchase.get("link", "")
    
    if not purchase_number or not link:
        return []
    
    logger.info(f"      🔍 Закупка №{purchase_number}: получение лотов...")
    
    # Получаем страницу закупки
    response = safe_request(link)
    if not response:
        logger.warning(f"      ⚠️  Не удалось загрузить страницу закупки")
        return []
    
    # Ищем noticeGuid на странице
    notice_guid = None
    soup = BeautifulSoup(response.text, "html.parser")
    
    # Ищем в ссылках на лоты
    lot_link = soup.find('a', href=re.compile(r'lot-list\.html|lot-info\.html|noticeGuid='))
    if lot_link:
        href = lot_link.get('href', '')
        guid_match = re.search(r'noticeGuid=([a-f0-9-]+)', href, re.I)
        if guid_match:
            notice_guid = guid_match.group(1)
    
    # Если не нашли - ищем в JavaScript
    if not notice_guid:
        page_text = response.text
        guid_match = re.search(r'noticeGuid["\s:=]+["\']?([a-f0-9-]+)["\']?', page_text, re.I)
        if guid_match:
            notice_guid = guid_match.group(1)
    
    if not notice_guid:
        logger.warning(f"      ⚠️  noticeGuid не найден")
        return []
    
    logger.info(f"         📋 noticeGuid: {notice_guid[:8]}...")
    
    # Получаем список лотов
    lot_list_url = f"{BASE_URL}/epz/order/notice/notice223/lot-list.html?purchaseNoticeNumber={purchase_number}&noticeGuid={notice_guid}"
    if "44" in fz_type:
        lot_list_url = f"{BASE_URL}/epz/order/notice/notice44/lot-list.html?purchaseNoticeNumber={purchase_number}&noticeGuid={notice_guid}"
    
    time.sleep(DELAY_BETWEEN_LOTS)
    lot_response = safe_request(lot_list_url)
    if not lot_response:
        logger.warning(f"      ⚠️  Не удалось загрузить список лотов")
        return []
    
    lots = parse_lot_list(lot_response.text, purchase_number, fz_type)
    logger.info(f"         📦 Найдено лотов: {len(lots)}")
    
    # Для каждого лота получаем детальную информацию
    lot_details = []
    for idx, lot in enumerate(lots, 1):
        logger.info(f"         🔍 Лот {idx}/{len(lots)}: {lot['lot_guid'][:8]}...")
        
        lot_info_url = get_lot_info_url(
            lot['lot_guid'], 
            notice_guid, 
            purchase_number, 
            fz_type
        )
        
        if not lot_info_url:
            continue
        
        time.sleep(DELAY_BETWEEN_LOTS)
        lot_info_response = safe_request(lot_info_url)
        
        if lot_info_response:
            lot_info = parse_lot_info(lot_info_response.text)
            lot_info['lot_guid'] = lot['lot_guid']
            lot_info['lot_number'] = lot.get('lot_number', '')
            lot_info['purchase_number'] = purchase_number
            lot_info['notice_guid'] = notice_guid
            lot_details.append(lot_info)
            logger.info(f"            ✅ Полей: {len(lot_info)}")
        else:
            logger.warning(f"            ⚠️  Не удалось загрузить лот")
    
    return lot_details


# ===== СОХРАНЕНИЕ В EXCEL =====

def save_to_excel(results, lot_details_all, filename, script_start_time, script_name, inn_names):
    """Сохранение результатов в Excel с расширенными данными"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # ===== ЛИСТ 1: lots (основная информация) =====
    ws_lots = wb.create_sheet(title="lots")
    
    headers = [
        "ID", "Номер закупки", "Название", "Организация", "Цена (руб.)",
        "Стадия", "ФЗ", "Тип", "Дата размещения", "Дата обновления",
        "Дата окончания", "Ссылка",
        # Новые колонки из деталей лотов (берём данные первого лота)
        "Номер плана", "Номер позиции плана", "Период планирования",
        "Начальная цена лота", "Место поставки (субъект)", "Место поставки (адрес)"
    ]
    
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws_lots.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    row = 2
    for inn, purchases in results.items():
        for purchase in purchases:
            # Основные данные закупки
            ws_lots.cell(row=row, column=1, value=row-1)
            ws_lots.cell(row=row, column=2, value=purchase.get("number", ""))
            ws_lots.cell(row=row, column=3, value=purchase.get("name", ""))
            ws_lots.cell(row=row, column=4, value=purchase.get("agency", ""))
            
            price_cell = ws_lots.cell(row=row, column=5, value=purchase.get("price", 0))
            price_cell.number_format = '#,##0.00₽'
            
            ws_lots.cell(row=row, column=6, value=purchase.get("stage", ""))
            ws_lots.cell(row=row, column=7, value=purchase.get("fz", ""))
            ws_lots.cell(row=row, column=8, value=purchase.get("type", ""))
            
            if purchase.get("published_date"):
                ws_lots.cell(row=row, column=9, value=purchase.get("published_date"))
            if purchase.get("updated_date"):
                ws_lots.cell(row=row, column=10, value=purchase.get("updated_date"))
            if purchase.get("end_date"):
                ws_lots.cell(row=row, column=11, value=purchase.get("end_date"))
            
            ws_lots.cell(row=row, column=12, value=purchase.get("link", ""))
            
            # Данные из лотов (берём первый лот)
            purchase_number = purchase.get("number", "")
            first_lot = next((l for l in lot_details_all if l.get('purchase_number') == purchase_number), None)
            
            if first_lot:
                ws_lots.cell(row=row, column=13, value=first_lot.get('plan_number', ''))
                ws_lots.cell(row=row, column=14, value=first_lot.get('plan_position_number', ''))
                ws_lots.cell(row=row, column=15, value=first_lot.get('planning_period', ''))
                
                if first_lot.get('initial_price'):
                    price_cell = ws_lots.cell(row=row, column=16, value=first_lot.get('initial_price'))
                    price_cell.number_format = '#,##0.00₽'
                
                ws_lots.cell(row=row, column=17, value=first_lot.get('delivery_region', ''))
                ws_lots.cell(row=row, column=18, value=first_lot.get('delivery_address', ''))
            
            row += 1
    
    # Автоширина
    for col in ws_lots.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws_lots.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # ===== ЛИСТ 2: lot_details (детали всех лотов) =====
    if lot_details_all:
        ws_details = wb.create_sheet(title="lot_details")
        
        detail_headers = [
            "№", "Номер закупки", "GUID лота", "Номер лота",
            "Наименование предмета договора", "Номер плана",
            "Номер позиции плана", "Начальная цена лота",
            "Период планирования", "Дата начала планирования",
            "Дата окончания планирования", "Размещено", "Обновлено",
            "Место поставки (субъект)", "Место поставки (адрес)"
        ]
        
        for col, header in enumerate(detail_headers, 1):
            cell = ws_details.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        for idx, lot in enumerate(lot_details_all, 2):
            ws_details.cell(row=idx, column=1, value=idx-1)
            ws_details.cell(row=idx, column=2, value=lot.get('purchase_number', ''))
            ws_details.cell(row=idx, column=3, value=lot.get('lot_guid', ''))
            ws_details.cell(row=idx, column=4, value=lot.get('lot_number', ''))
            ws_details.cell(row=idx, column=5, value=lot.get('lot_name', ''))
            ws_details.cell(row=idx, column=6, value=lot.get('plan_number', ''))
            ws_details.cell(row=idx, column=7, value=lot.get('plan_position_number', ''))
            
            if lot.get('initial_price'):
                price_cell = ws_details.cell(row=idx, column=8, value=lot.get('initial_price'))
                price_cell.number_format = '#,##0.00₽'
            
            ws_details.cell(row=idx, column=9, value=lot.get('planning_period', ''))
            
            if lot.get('planning_start_date'):
                ws_details.cell(row=idx, column=10, value=lot.get('planning_start_date'))
            if lot.get('planning_end_date'):
                ws_details.cell(row=idx, column=11, value=lot.get('planning_end_date'))
            if lot.get('published_date'):
                ws_details.cell(row=idx, column=12, value=lot.get('published_date'))
            if lot.get('updated_date'):
                ws_details.cell(row=idx, column=13, value=lot.get('updated_date'))
            
            ws_details.cell(row=idx, column=14, value=lot.get('delivery_region', ''))
            ws_details.cell(row=idx, column=15, value=lot.get('delivery_address', ''))
        
        # Автоширина
        for col in ws_details.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws_details.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # ===== ЛИСТ 3: log =====
    ws_log = wb.create_sheet(title="log")
    
    log_headers = ["Date time", "Script name", "Script mod time", "Command line", "Stage", "INN list", "Years"]
    for col, header in enumerate(log_headers, 1):
        cell = ws_log.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    script_mod_time = datetime.fromtimestamp(os.path.getmtime(__file__))
    inn_list = list(results.keys())
    
    ws_log.cell(row=2, column=1, value=script_start_time)
    ws_log.cell(row=2, column=2, value=script_name)
    ws_log.cell(row=2, column=3, value=script_mod_time)
    ws_log.cell(row=2, column=4, value=json.dumps(sys.argv))
    ws_log.cell(row=2, column=5, value=1)
    ws_log.cell(row=2, column=6, value=json.dumps(inn_list))
    ws_log.cell(row=2, column=7, value="[]")
    
    for col in ws_log.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws_log.column_dimensions[col_letter].width = min(max_length + 2, 30)
    
    wb.save(filename)
    return True


# ===== ГЛАВНАЯ ФУНКЦИЯ =====

def main():
    script_start_time = datetime.now()
    script_name = os.path.basename(__file__)
    
    logger.info("=" * 70)
    logger.info("РАСШИРЕННЫЙ ПАРСЕР ЗАКУПОК")
    logger.info(f"Время запуска: {script_start_time.strftime('%d.%m.%Y %H:%M:%S')}")
    logger.info("=" * 70)
    
    args = parse_arguments()
    date_from, date_to = get_date_range(args)
    
    if args.inn:
        inn_dict = {inn: inn for inn in args.inn}
        logger.info(f"ИНН из командной строки: {len(inn_dict)}")
    else:
        inn_dict = load_inn_from_file()
        if not inn_dict:
            logger.error("❌ Нет ИНН для поиска")
            return
        logger.info(f"ИНН из inn_list.txt: {len(inn_dict)}")
    
    logger.info(f"Период: {date_from.strftime('%d.%m.%Y')} - {date_to.strftime('%d.%m.%Y')}")
    logger.info(f"Сбор деталей лотов: {'НЕТ' if args.skip_lots else 'ДА'}")
    
    results = {}
    all_lot_details = []
    
    for idx, (inn, name) in enumerate(inn_dict.items(), 1):
        logger.info(f"\n{'='*70}")
        logger.info(f"[{idx}/{len(inn_dict)}] 🏢 {name} (ИНН: {inn})")
        logger.info(f"{'='*70}")
        
        # Шаг 1: Поиск закупок
        purchases = search_purchases(
            inn, date_from, date_to,
            args.page_size, args.max_pages, args.delay
        )
        results[inn] = purchases
        logger.info(f"   📊 Найдено закупок: {len(purchases)}")
        
        # Шаг 2: Сбор деталей лотов
        if not args.skip_lots:
            for purchase in purchases:
                try:
                    lot_details = fetch_purchase_lot_details(purchase, args.skip_lots)
                    all_lot_details.extend(lot_details)
                    logger.info(f"      ✅ Лотов: {len(lot_details)}")
                except Exception as e:
                    logger.error(f"      ❌ Ошибка сбора лотов: {e}")
                
                time.sleep(DELAY_BETWEEN_PURCHASES)
        
        if idx < len(inn_dict):
            logger.info(f"\n⏳ Пауза {DELAY_BETWEEN_INN}с...")
            time.sleep(DELAY_BETWEEN_INN)
    
    # Сохранение результатов
    total_purchases = sum(len(p) for p in results.values())
    
    if total_purchases > 0:
        folder_path, folder_number = get_output_folder()
        date_str = datetime.now().strftime("%d.%m.%Y")
        filename = os.path.join(folder_path, f"zakupki_extended_{date_str}_{folder_number:02d}.xlsx")
        
        save_to_excel(results, all_lot_details, filename, script_start_time, script_name, inn_dict)
        
        logger.info(f"\n{'='*70}")
        logger.info(f"📊 ИТОГОВАЯ СТАТИСТИКА")
        logger.info(f"{'='*70}")
        logger.info(f"   Начало: {script_start_time.strftime('%d.%m.%Y %H:%M:%S')}")
        logger.info(f"   Конец:  {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
        logger.info(f"   Всего закупок: {total_purchases}")
        logger.info(f"   Всего лотов: {len(all_lot_details)}")
        logger.info(f"   💾 Файл: {filename}")
        logger.info(f"{'='*70}")
    else:
        logger.info("\n❌ Закупок не найдено")


if __name__ == "__main__":
    main()

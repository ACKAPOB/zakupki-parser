#!/usr/bin/env python3
"""
Финальный парсер v6:
- Конфигурация вынесена в config.py
- Результаты сохраняются в папку output/ДД.ММ.ГГГГ_NN/
- Повторные попытки при сбое
- Отчёт о пропущенных планах
"""

import sys
import os
import re
import time
import requests
import logging
import glob
from datetime import datetime
from pathlib import Path
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright, TimeoutError
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import argparse

# Импортируем конфиг
import config

# Настройка логирования (будет перенаправлено в папку результата)
logger = logging.getLogger(__name__)


def setup_logging(log_file):
    """Настройка логирования в файл и консоль"""
    logger.setLevel(logging.INFO)
    
    # Очистить старые хендлеры
    logger.handlers.clear()
    
    # Файловый хендлер
    file_handler = logging.FileHandler(log_file, encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    logger.addHandler(file_handler)
    
    # Консольный хендлер
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    logger.addHandler(console_handler)


def parse_arguments():
    parser = argparse.ArgumentParser(description="Парсер планов-графиков v6")
    parser.add_argument("-i", "--inn", help="ИНН организации")
    parser.add_argument("-f", "--file", help=f"Файл со списком ИНН (по умолчанию: {config.INN_LIST_FILE})")
    parser.add_argument("--year", type=int, default=config.DEFAULT_YEAR, help="Год плана")
    parser.add_argument("--fz", choices=["44", "223", "all"], default=config.DEFAULT_FZ, help="ФЗ")
    parser.add_argument("--no-headless", action="store_true", help="Браузер с интерфейсом")
    parser.add_argument("--max-pages", type=int, default=config.MAX_PAGINATION_PAGES, help="Максимум страниц пагинации")
    parser.add_argument("--skip-details", action="store_true", help="Не собирать детальные данные")
    return parser.parse_args()


def load_inn_list(filename):
    inn_list = []
    if not os.path.exists(filename):
        logger.error(f"❌ Файл не найден: {filename}")
        sys.exit(1)
    with open(filename, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line: continue
            if '|' in line:
                parts = line.split('|', 1)
                inn, name = parts[0].strip(), parts[1].strip()
            else:
                inn, name = line.strip(), ""
            if inn.isdigit() and len(inn) in [10, 12]:
                inn_list.append({'inn': inn, 'name': name})
    return inn_list


def get_output_folder():
    """Создать папку output/ДД.ММ.ГГГГ_NN/ и вернуть путь"""
    # Проверяем переменную окружения от run_all.py
    session_dir = os.environ.get('ZAKUPKI_OUTPUT_DIR')
    
    if session_dir and os.path.isdir(session_dir):
        folder_path = Path(session_dir)
        logger.info(f"📁 Используем папку сессии: {folder_path}")
        return folder_path, 0  # Номер не важен
    
    # Обычная логика для отдельного запуска
    today = datetime.now().strftime("%d.%m.%Y")
    
    pattern = f"{config.OUTPUT_DIR}/{today}_*"
    existing_folders = glob.glob(pattern)
    
    if not existing_folders:
        folder_number = 1
    else:
        numbers = []
        for folder in existing_folders:
            match = re.search(r'(\d{2}\.\d{2}\.\d{4})_(\d+)', folder)
            if match:
                numbers.append(int(match.group(2)))
        folder_number = max(numbers) + 1 if numbers else 1
    
    folder_name = f"{today}_{folder_number:02d}"
    folder_path = config.OUTPUT_DIR / folder_name
    folder_path.mkdir(parents=True, exist_ok=True)
    
    logger.info(f"📁 Папка результата: {folder_path}")
    return folder_path, folder_number


def get_file_number(folder_path):
    """Определить номер файла внутри папки"""
    pattern = f"{folder_path}/Plans_*.xlsx"
    existing_files = glob.glob(pattern)
    
    if not existing_files:
        return 1
    
    numbers = []
    for f in existing_files:
        match = re.search(r'Plans_\d{2}\.\d{2}\.\d{4}_(\d+)\.xlsx', f)
        if match:
            numbers.append(int(match.group(1)))
    
    return max(numbers) + 1 if numbers else 1


def build_search_params(search_query, year, fz_type, page, page_size):
    params = {
        "morphology": "on", "search-filter": "Дате размещения",
        "structured": "true", "customerPlaceWithNested": "on",
        "sortBy": "BY_MODIFY_DATE", "sortDirection": "false",
        "pageNumber": str(page), "recordsPerPage": f"_{page_size}",
        "searchString": search_query,
    }
    if fz_type in ("44", "all"): params["fz44"] = "on"
    if fz_type in ("223", "all"): params["fz223"] = "on"
    if year: params["actualPeriodRangeYearFrom"] = str(year)
    return params


def parse_date(date_text):
    if not date_text: return None
    try: return datetime.strptime(date_text.strip(), "%d.%m.%Y").date()
    except: return None


def extract_plan_info(block):
    try:
        record = {}
        header = block.find("div", class_="registry-entry__header-top__title")
        if header:
            values = [e.strip() for e in header.text.split("\n") if e.strip()]
            if len(values) >= 2:
                record["fz"] = values[0]
                record["type"] = values[1]
        number_block = block.find("div", class_="registry-entry__header-mid__number")
        if number_block:
            link_tag = number_block.find("a")
            if link_tag:
                href = link_tag.get("href", "")
                record["link"] = config.BASE_URL + href
                guid_match = re.search(r'guid=([a-f0-9-]+)', href, re.I)
                if guid_match: record["guid"] = guid_match.group(1)
                number_text = link_tag.text.strip()
                match = re.search(r"№\s*([^\s]+)", number_text)
                if match: record["number"] = match.group(1)
        body = block.find("div", class_="registry-entry__body")
        if body:
            name_block = body.find("div", class_="registry-entry__body-value")
            if name_block: record["name"] = re.sub(r"\s+", " ", name_block.text.strip())
            href_block = body.find("div", class_="registry-entry__body-href")
            if href_block:
                link_tag = href_block.find("a")
                if link_tag: record["organization"] = link_tag.text.strip()
        right_block = block.find("div", class_="registry-entry__right-block")
        if right_block:
            data_block = right_block.find("div", class_="data-block")
            if data_block:
                for item in data_block.find_all("div", class_="data-block__item"):
                    title = item.find("div", class_="data-block__title")
                    value = item.find("div", class_="data-block__value")
                    if title and value:
                        title_text = title.text.strip()
                        if title_text == "Размещено": record["published_date"] = parse_date(value.text)
                        elif title_text == "Обновлено": record["updated_date"] = parse_date(value.text)
        return record if record.get("number") else None
    except Exception as e:
        return None


def search_plans(search_query, year, fz_type="all", page_size=None, max_pages=None):
    if page_size is None: page_size = config.SEARCH_PAGE_SIZE
    if max_pages is None: max_pages = config.MAX_SEARCH_PAGES
    
    all_plans, page, found = [], 1, set()
    logger.info(f"\n🔍 🔍 🔍 Поиск планов по: {search_query}")
    while page <= max_pages:
        params = build_search_params(search_query, year, fz_type, page, page_size)
        try:
            print(f"   Страница {page}...", end=" ", flush=True)
            response = requests.get(config.SEARCH_URL, headers=config.HEADERS, params=params, timeout=30)
            if response.status_code != 200:
                print(f"Ошибка HTTP {response.status_code}")
                break
            soup = BeautifulSoup(response.text, "html.parser")
            if page == 1:
                total_block = soup.find("div", class_="search-results__total")
                if total_block: print(f"найдено: {total_block.text.strip()}")
            blocks = soup.find_all("div", class_="search-registry-entry-block")
            if not blocks:
                print("записей не найдено")
                break
            page_plans = []
            for block in blocks:
                plan = extract_plan_info(block)
                if plan and plan.get("number") not in found:
                    found.add(plan.get("number"))
                    page_plans.append(plan)
            print(f"записей: {len(page_plans)}")
            all_plans.extend(page_plans)
            pagination = soup.find("div", class_="search-results__pagination")
            next_link = None
            if pagination: next_link = pagination.find("a", title="Следующая страница")
            if not next_link:
                print("   Достигнут конец списка")
                break
            page += 1
            if page <= max_pages:
                time.sleep(config.DELAY_BETWEEN_SEARCH_PAGES)
        except Exception as e:
            print(f"Ошибка: {e}")
            break
    return all_plans


def close_modal(page):
    try:
        page.evaluate('''
            () => {
                const modal = document.querySelector('#modal-customer');
                const overlay = document.querySelector('.popupModalOverlay');
                if (modal) modal.remove();
                if (overlay) overlay.remove();
            }
        ''')
        page.wait_for_timeout(300)
    except: pass


def get_pagination_info(page):
    try:
        active = page.query_selector('.page__link_active span.link-text')
        current = 1
        if active:
            text = active.text_content().strip()
            if text.isdigit(): current = int(text)
        links = page.query_selector_all('a.page__link span.link-text')
        max_page = 1
        for link in links:
            text = link.text_content().strip()
            if text.isdigit():
                num = int(text)
                if num > max_page: max_page = num
        return current, max_page
    except: return 1, 1


def click_page(page, page_num):
    try:
        close_modal(page)
        page.evaluate(f'_searchPurchasegoToPage({page_num})')
        page.wait_for_load_state('networkidle')
        page.wait_for_timeout(1500)
        close_modal(page)
        return True
    except Exception as e:
        logger.error(f"❌ Ошибка перехода на страницу {page_num}: {e}")
        return False


def parse_placement(placement_text):
    result = {'Срок размещения': '', 'Способ закупки': ''}
    if not placement_text:
        return result
    
    term_match = re.search(r'Срок размещения:\s*([^\,]+)', placement_text)
    if term_match:
        term_text = term_match.group(1).strip()
        for month_name, month_num in config.MONTHS.items():
            if month_name.lower() in term_text.lower():
                year_match = re.search(r'(\d{4})', term_text)
                if year_match:
                    year = year_match.group(1)
                    result['Срок размещения'] = f"{month_num}.{year}"
                break
        if not result['Срок размещения']:
            result['Срок размещения'] = term_text
    
    method_match = re.search(r'Способ закупки:\s*(.+?)(?:,|$)', placement_text)
    if method_match:
        result['Способ закупки'] = method_match.group(1).strip()
    
    return result


def fetch_position_details(context, position_link):
    if not position_link:
        return {}
    
    detail_page = None
    try:
        detail_page = context.new_page()
        detail_page.goto(position_link, wait_until='networkidle', timeout=30000)
        detail_page.wait_for_timeout(1000)
        
        details = {}
        
        title = detail_page.query_selector('h1, h2, .registry-entry__header-mid__number')
        if title:
            details['title'] = title.text_content().strip()
        
        info_blocks = detail_page.query_selector_all('.registry-entry__body-value, .data-block__value')
        for block in info_blocks:
            text = block.text_content().strip()
            if text and len(text) > 10:
                parent = block.query_selector('xpath=ancestor::div[contains(@class, "data-block__item")]')
                if parent:
                    label = parent.query_selector('.data-block__title')
                    if label:
                        label_text = label.text_content().strip().lower()
                        details[label_text] = text
        
        tables = detail_page.query_selector_all('table')
        for table in tables:
            rows = table.query_selector_all('tbody tr')
            for row in rows:
                cells = row.query_selector_all('td')
                if len(cells) >= 2:
                    key = cells[0].text_content().strip().lower()
                    value = cells[1].text_content().strip()
                    if key and value:
                        details[key] = value
        
        page_text = detail_page.content()
        
        patterns = [
            (r'ОКПД2[:\s]+([^\n<]+)', 'ОКПД2'),
            (r'КТРУ[:\s]+([^\n<]+)', 'КТРУ'),
            (r'(?:единица|ед\.? измерения|едизм)[:\s]+([^\n<]+)', 'Единица измерения'),
            (r'(?:количество|объем)[:\s]+([^\n<]+)', 'Количество'),
            (r'(?:срок.*?(?:поставки|исполнения|оказания))[:\s]+([^\n<]+)', 'Срок'),
            (r'(?:место.*?(?:поставки|оказания|выполнения))[:\s]+([^\n<]+)', 'Место'),
            (r'(?:описание|наименование объекта)[\s:]+([^\n]{50,})', 'Описание'),
        ]
        
        for pattern, key in patterns:
            match = re.search(pattern, page_text, re.I)
            if match:
                value = match.group(1).strip()
                if key == 'Описание':
                    value = value[:200]
                details[key] = value
        
        return details
        
    except TimeoutError:
        return {'_error': 'Таймаут загрузки'}
    except Exception as e:
        return {'_error': str(e)[:100]}
    finally:
        if detail_page:
            try:
                detail_page.close()
            except:
                pass


def parse_positions_from_table(page, table_element):
    positions = []
    rows = table_element.query_selector_all('tbody tr')
    
    for i, row in enumerate(rows):
        try:
            cols = row.query_selector_all('td')
            if len(cols) < 2: continue
            
            number = cols[1].text_content().strip()
            if not number or not number.isdigit(): continue
            
            name_col = cols[2] if len(cols) > 2 else None
            name = ""
            position_link = ""
            
            if name_col:
                link = name_col.query_selector('a[href*="tru-plan/card/common-info.html"]')
                if link:
                    href = link.get_attribute('href')
                    if href:
                        position_link = config.BASE_URL + href
                    name = link.text_content().strip()
                else:
                    name = name_col.text_content().strip()
            
            price_col = cols[3] if len(cols) > 3 else None
            price_text = price_col.text_content().strip() if price_col else "0"
            price = parse_price(price_text)
            
            placement_col = cols[4] if len(cols) > 4 else None
            placement_text = placement_col.text_content().strip() if placement_col else ""
            placement_parsed = parse_placement(placement_text)
            
            term_col = cols[5] if len(cols) > 5 else None
            term = term_col.text_content().strip() if term_col else ""
            
            customer = ""
            if i + 1 < len(rows):
                next_row = rows[i + 1]
                next_cols = next_row.query_selector_all('td')
                if len(next_cols) >= 3:
                    second_col = next_cols[1].text_content().strip()
                    if "заказчик" in second_col.lower():
                        customer = next_cols[2].text_content().strip()
            
            positions.append({
                "number": number, 
                "name": name, 
                "price": price,
                "Срок размещения": placement_parsed['Срок размещения'],
                "Способ закупки": placement_parsed['Способ закупки'],
                "term": term, 
                "customer": customer,
                "position_link": position_link,
                "details": {}
            })
        except Exception as e:
            continue
    
    return positions


def parse_price(price_text):
    if not price_text: return 0.0
    cleaned = re.sub(r"[₽\s]", "", price_text)
    cleaned = cleaned.replace(",", ".").replace(chr(160), "")
    try: return float(cleaned)
    except: return 0.0


def extract_positions_from_plan(context, page, plan, max_pagination_pages=None, skip_details=False):
    """Извлечение позиций с повторными попытками"""
    if max_pagination_pages is None: max_pagination_pages = config.MAX_PAGINATION_PAGES
    
    guid = plan.get('guid')
    reg_number = plan.get('number')
    
    if not guid:
        logger.warning(f"⚠️  GUID не найден")
        return [], []
    
    logger.info(f"    План {reg_number} (GUID: {guid[:8]}...)")
    url = f"{config.BASE_URL}/epz/orderplan/purchase-plan/card/position-info.html?guid={guid}"
    
    # Повторные попытки загрузки
    for attempt in range(1, config.MAX_RETRIES + 1):
        try:
            if attempt > 1:
                logger.info(f"      🔄  🔄 Попытка {attempt}/{config.MAX_RETRIES}...")
                time.sleep(config.DELAY_RETRY)
            
            page.goto(url, wait_until='networkidle', timeout=config.PAGE_LOAD_TIMEOUT)
            page.wait_for_selector('table', timeout=config.PAGE_LOAD_TIMEOUT)
            break  # Успешно загрузились
        except TimeoutError:
            if attempt == config.MAX_RETRIES:
                logger.error(f"      ❌❌ Таблица не загрузилась после {config.MAX_RETRIES} попыток")
                return [], []
            else:
                logger.warning(f"      ⏱️  ⏱️  Таймаут, пробуем ещё...")
        except Exception as e:
            if attempt == config.MAX_RETRIES:
                logger.error(f"      ❌ Ошибка загрузки: {e}")
                return [], []
            else:
                logger.warning(f"      ️  ⚠️  Ошибка: {e}, пробуем ещё...")
    
    close_modal(page)
    
    all_positions = []
    seen_keys = set()
    errors = []
    
    current_page, max_page = get_pagination_info(page)
    logger.info(f"       Всего страниц пагинации: {max_page}")
    
    if max_page > max_pagination_pages:
        logger.warning(f"⚠️  Слишком много страниц ({max_page}), ограничиваем до {max_pagination_pages}")
        max_page = max_pagination_pages
    
    long_term_processed = False
    
    for page_num in range(1, max_page + 1):
        logger.info(f"\n       Страница {page_num}/{max_page}")
        
        if page_num > 1:
            if not click_page(page, page_num):
                logger.error(f"         ❌ Не удалось перейти")
                break
            page.wait_for_timeout(500)
        
        try:
            tables = page.query_selector_all('table')
            logger.info(f"          📊 📊 Найдено таблиц: {len(tables)}")
        except Exception as e:
            logger.error(f"         ❌ Ошибка получения таблиц: {e}")
            break
        
        for table_idx, table in enumerate(tables):
            table_type = "Долгосрочная" if table_idx == 1 else "Основная"
            
            if table_type == "Долгосрочная" and page_num > 1:
                if long_term_processed:
                    continue
            
            try:
                positions = parse_positions_from_table(page, table)
            except Exception as e:
                logger.error(f"          Ошибка парсинга таблицы: {e}")
                continue
            
            new_count = 0
            
            for pos in positions:
                try:
                    key = f"{table_type}|{pos['number']}"
                    if key in seen_keys:
                        continue
                    
                    seen_keys.add(key)
                    pos['table_type'] = table_type
                    
                    if not skip_details and pos.get('position_link'):
                        logger.info(f"             🔍 🔍 Позиция #{pos['number']}: сбор деталей...")
                        try:
                            pos['details'] = fetch_position_details(context, pos['position_link'])
                            if '_error' in pos['details']:
                                errors.append(f"🔍 Позиция #{pos['number']}: {pos['details']['_error']}")
                        except Exception as e:
                            errors.append(f"🔍 Позиция #{pos['number']}: {str(e)[:100]}")
                            pos['details'] = {'_error': str(e)[:100]}
                        
                        time.sleep(config.DELAY_AFTER_POSITION)
                    
                    all_positions.append(pos)
                    new_count += 1
                except Exception as e:
                    errors.append(f"🔍 Позиция #{pos.get('number', '?')}: {str(e)[:100]}")
                    continue
            
            if table_type == "Долгосрочная":
                long_term_processed = True
            
            logger.info(f"          Таблица {table_idx + 1} ({table_type}): "
                       f"{len(positions)} позиций, новых: {new_count}")
        
        if page_num < max_page:
            time.sleep(config.DELAY_BETWEEN_PAGES)
    
    all_positions.sort(key=lambda x: int(x['number']))
    logger.info(f"\n       Всего уникальных позиций: {len(all_positions)}")
    
    if errors:
        logger.warning(f"      ⚠️  Ошибок при сборе: {len(errors)}")
    
    return all_positions, errors


def save_to_excel(all_data, all_errors, filename, log_file):
    logger.info(f"\n💾 💾 💾 Сохранение в {filename}...")
    
    all_detail_fields = set()
    for item in all_data:
        for pos in item['positions']:
            if 'details' in pos:
                all_detail_fields.update(pos['details'].keys())
    all_detail_fields.discard('_error')
    
    wb = openpyxl.Workbook()
    
    ws = wb.active
    ws.title = "Позиции планов"
    
    main_headers = ["ИНН", "Организация", "№ плана", "Тип ФЗ", "Тип позиции", "№ позиции",
                   "Наименование", "Цена (руб.)", "Срок размещения", "Способ закупки", 
                   "Срок исполнения", "Заказчик"]
    
    detail_headers = sorted(list(all_detail_fields))
    all_headers = main_headers + detail_headers
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    row_num = 2
    for item in all_data:
        for pos in item['positions']:
            ws.cell(row=row_num, column=1, value=item['inn'])
            ws.cell(row=row_num, column=2, value=item['organization'])
            ws.cell(row=row_num, column=3, value=item['plan_number'])
            ws.cell(row=row_num, column=4, value=item['law_type'])
            ws.cell(row=row_num, column=5, value=pos.get('table_type', 'Основная'))
            ws.cell(row=row_num, column=6, value=pos['number'])
            ws.cell(row=row_num, column=7, value=pos['name'])
            ws.cell(row=row_num, column=8, value=pos['price'])
            ws.cell(row=row_num, column=9, value=pos.get('Срок размещения', ''))
            ws.cell(row=row_num, column=10, value=pos.get('Способ закупки', ''))
            ws.cell(row=row_num, column=11, value=pos['term'])
            ws.cell(row=row_num, column=12, value=pos['customer'])
            
            details = pos.get('details', {})
            for col_offset, field in enumerate(detail_headers, 13):
                value = details.get(field, '')
                ws.cell(row=row_num, column=col_offset, value=value)
            
            row_num += 1
    
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value: max_length = max(max_length, len(str(cell.value)))
            except: pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 80)
    
    total_row = row_num + 1
    ws.cell(row=total_row, column=1, value=f"Всего позиций: {row_num - 2}")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    
    # Лист лога
    ws_log = wb.create_sheet(title="Лог работы")
    
    log_headers = ["Время", "Уровень", "Сообщение"]
    for col, header in enumerate(log_headers, 1):
        cell = ws_log.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    log_rows = 2
    if os.path.exists(log_file):
        with open(log_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line: continue
                match = re.match(r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d+) - (\w+) - (.+)', line)
                if match:
                    ws_log.cell(row=log_rows, column=1, value=match.group(1))
                    ws_log.cell(row=log_rows, column=2, value=match.group(2))
                    ws_log.cell(row=log_rows, column=3, value=match.group(3))
                    log_rows += 1
    
    ws_log.column_dimensions['A'].width = 22
    ws_log.column_dimensions['B'].width = 12
    ws_log.column_dimensions['C'].width = 120
    
    # Лист ошибок
    if all_errors:
        ws_err = wb.create_sheet(title="Ошибки")
        err_headers = ["№", "Описание ошибки"]
        for col, header in enumerate(err_headers, 1):
            cell = ws_err.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for idx, err in enumerate(all_errors, 2):
            ws_err.cell(row=idx, column=1, value=idx - 1)
            ws_err.cell(row=idx, column=2, value=err)
        
        ws_err.column_dimensions['A'].width = 5
        ws_err.column_dimensions['B'].width = 100
    
    wb.save(filename)
    logger.info(f"✅ ✅ ✅ Сохранено: {filename}")
    return filename


def main():
    start_time = datetime.now()
    
    args = parse_arguments()
    
    # Создаём папку результата
    output_folder, folder_number = get_output_folder()
    
    # Настраиваем логирование в папку результата
    log_file = output_folder / config.LOG_FILE_NAME
    setup_logging(log_file)
    
    logger.info("=" * 70)
    logger.info("ПАРСЕР ПЛАНОВ-ГРАФИКОВ v6")
    logger.info(f"Начало работы: {start_time.strftime('%d.%m.%Y %H:%M:%S')}")
    logger.info(f"📁 📁 Папка результата: {output_folder}")
    logger.info("=" * 70)
    
    # Определяем список ИНН
    inn_file = args.file if args.file else config.INN_LIST_FILE
    if args.inn:
        inn_list = [{'inn': args.inn, 'name': f'ИНН {args.inn}'}]
    else:
        inn_list = load_inn_list(inn_file)
    
    skip_details = args.skip_details or not config.COLLECT_DETAILS
    
    logger.info(f"ИНН для обработки: {len(inn_list)}")
    logger.info(f"Год: {args.year}")
    logger.info(f"Максимум страниц пагинации: {args.max_pages}")
    logger.info(f"Сбор деталей: {'НЕТ' if skip_details else 'ДА'}")
    logger.info(f"Повторные попытки: {config.MAX_RETRIES}")
    logger.info("=" * 70)
    
    all_data = []
    all_errors = []
    failed_plans = []
    stats = {
        'inn_processed': 0,
        'plans_found': 0,
        'positions_total': 0,
        'errors_total': 0
    }
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=(not args.no_headless and config.HEADLESS))
        context = browser.new_context(viewport={"width": config.BROWSER_WIDTH, "height": config.BROWSER_HEIGHT}, ignore_https_errors=True)
        page = context.new_page()
        
        for idx, item in enumerate(inn_list, 1):
            inn = item['inn']
            org_name = item['name']
            
            logger.info(f"\n{'='*70}")
            logger.info(f"[{idx}/{len(inn_list)}]  {org_name} (ИНН: {inn})")
            logger.info(f"{'='*70}")
            
            try:
                plans = search_plans(
                    search_query=inn, year=args.year, fz_type=args.fz,
                    page_size=config.SEARCH_PAGE_SIZE, max_pages=config.MAX_SEARCH_PAGES
                )
            except Exception as e:
                logger.error(f"❌ Ошибка поиска планов: {e}")
                all_errors.append(f"ИНН {inn}: ошибка поиска - {str(e)[:100]}")
                continue
            
            if not plans:
                logger.warning(f"⚠️  Планы не найдены")
                continue
            
            for plan in plans:
                try:
                    positions, errors = extract_positions_from_plan(
                        context, page, plan, args.max_pages, skip_details
                    )
                    if positions:
                        all_data.append({
                            'inn': inn, 'organization': org_name,
                            'plan_number': plan.get('number', ''),
                            'law_type': plan.get('fz', ''),
                            'positions': positions
                        })
                        stats['positions_total'] += len(positions)
                    else:
                        failed_plans.append({
                            'inn': inn,
                            'org': org_name,
                            'plan_number': plan.get('number', ''),
                            'guid': plan.get('guid', '')
                        })
                        logger.warning(f"      ⚠️  ⚠️  ⚠️  План {plan.get('number')} пропущен")
                    
                    all_errors.extend(errors)
                    stats['errors_total'] += len(errors)
                except Exception as e:
                    logger.error(f"❌ Ошибка обработки плана {plan.get('number')}: {e}")
                    all_errors.append(f"⚠️  План {plan.get('number')}: {str(e)[:100]}")
                    failed_plans.append({
                        'inn': inn,
                        'org': org_name,
                        'plan_number': plan.get('number', ''),
                        'guid': plan.get('guid', ''),
                        'error': str(e)[:100]
                    })
                
                time.sleep(config.DELAY_BETWEEN_PLANS)
            
            stats['plans_found'] += len(plans)
            stats['inn_processed'] += 1
            
            if idx < len(inn_list):
                logger.info(f"\n⏳ ⏳ ⏳ Пауза {config.DELAY_BETWEEN_INN}с перед следующей организацией...")
                time.sleep(config.DELAY_BETWEEN_INN)
        
        browser.close()
    
    end_time = datetime.now()
    duration = end_time - start_time
    
    # Сохраняем Excel
    file_number = get_file_number(output_folder)
    date_str = datetime.now().strftime("%d.%m.%Y")
    filename = output_folder / f"Plans_{date_str}_{file_number:02d}.xlsx"
    
    if all_data or all_errors or failed_plans:
        save_to_excel(all_data, all_errors, filename, log_file)
    
    # Финальная статистика
    logger.info(f"\n{'='*70}")
    logger.info("  📊 ИТОГОВАЯ СТАТИСТИКА")
    logger.info(f"{'='*70}")
    logger.info(f"   Начало: {start_time.strftime('%d.%m.%Y %H:%M:%S')}")
    logger.info(f"   Конец:  {end_time.strftime('%d.%m.%Y %H:%M:%S')}")
    logger.info(f"   Длительность: {duration}")
    logger.info(f"   Обработано ИНН: {stats['inn_processed']}/{len(inn_list)}")
    logger.info(f"   Найдено планов: {stats['plans_found']}")
    logger.info(f"   Всего позиций: {stats['positions_total']}")
    logger.info(f"   Ошибок: {stats['errors_total']}")
    logger.info(f"   Пропущено планов: {len(failed_plans)}")
    
    if failed_plans:
        logger.warning(f"\n️  ⚠️  ⚠️  СПИСОК ПРОПУЩЕННЫХ ПЛАНОВ:")
        for fp in failed_plans:
            error_info = f" ({fp['error']})" if 'error' in fp else ""
            logger.warning(f"      - ИНН {fp['inn']}, план {fp['plan_number']}{error_info}")
    
    logger.info(f"\n💾 💾 💾 Файл сохранён: {filename}")
    logger.info(f" 📁 📁 Папка: {output_folder}")
    logger.info(f"{'='*70}")


if __name__ == "__main__":
    main()
